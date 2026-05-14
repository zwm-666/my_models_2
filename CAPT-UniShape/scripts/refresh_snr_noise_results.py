"""Refresh SNR noise results and update the summary workbook.

The workbook's original no-noise result sheets are left untouched. The SNR
comparison sheet uses the clean metrics from the same trained models as the
noisy evaluations so that drop columns are protocol-aligned.
"""

from __future__ import annotations

import argparse
import csv
from copy import copy
import json
import math
import shutil
import sys
from pathlib import Path
from typing import Any, Iterable

import numpy as np
from openpyxl import load_workbook
from sklearn.decomposition import PCA
from sklearn.ensemble import RandomForestClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler
from sklearn.svm import LinearSVC, SVC
import yaml


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


CANONICAL_MODEL_ORDER = ["proposed", "logreg", "svm", "random_forest", "mlp", "cnn1d", "lstm", "transformer", "itransformer"]
DEFAULT_MODELS = ["proposed", "logreg", "svm", "random_forest", "mlp", "cnn1d", "transformer", "itransformer"]
DEFAULT_SNR_DBS = [40.0, 35.0, 30.0, 25.0, 20.0, 15.0, 10.0]
SUMMARY_FIELDNAMES = [
    "ratio",
    "model",
    "category",
    "snr_db",
    "snr_db_numeric",
    "actual_snr_db_mean",
    "noise_targets",
    "test_accuracy",
    "accuracy_drop",
    "test_macro_f1",
    "macro_f1_drop",
    "test_weighted_f1",
    "weighted_f1_drop",
    "test_inference_ms",
    "parameter_count",
    "data_path",
    "metrics_path",
    "clean_alignment_source",
]
COMPACT_SUMMARY_FIELDNAMES = [
    "model",
    "snr_db",
    "accuracy",
    "accuracy_drop",
    "macro_f1",
    "macro_f1_drop",
    "weighted_f1",
    "weighted_f1_drop",
    "data_path",
    "metrics_path",
    "clean_alignment_source",
]
EXPORT_FIELDNAMES = [
    "model",
    "snr_db",
    "test_accuracy",
    "accuracy_drop",
    "test_macro_f1",
    "macro_f1_drop",
    "test_weighted_f1",
    "weighted_f1_drop",
    "test_inference_ms",
    "parameter_count",
    "data_path",
    "metrics_path",
    "alignment_source",
]
MODEL_CATEGORIES = {
    "proposed": "proposed",
    "logreg": "traditional_ml",
    "svm": "traditional_ml",
    "random_forest": "traditional_ml",
    "mlp": "deep_learning",
    "cnn1d": "deep_learning",
    "lstm": "deep_learning",
    "transformer": "transformer",
    "itransformer": "itransformer",
}
MODEL_DISPLAY_NAMES = {
    "logreg": "Logistic Regression",
    "random_forest": "Random Forest",
    "mlp": "MLP",
    "cnn1d": "1D-CNN",
    "transformer": "Transformer",
    "itransformer": "iTransformer",
    "proposed": "所提模型",
    "svm": "SVM",
    "lstm": "LSTM",
}


def _resolve_path_string(value: Any) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    path = Path(raw)
    if path.is_absolute():
        return str(path)
    return str((ROOT / path).resolve())


def _pct_str_to_float(value: Any) -> float:
    text = str(value or "").strip()
    if text.endswith("%"):
        return float(text[:-1]) / 100.0
    return _safe_float(text)


def load_model_order_from_current_comparison_config(config_path: Path) -> list[str]:
    payload = yaml.safe_load(Path(config_path).read_text(encoding="utf-8")) or {}
    baselines = payload.get("baselines", {}) if isinstance(payload.get("baselines", {}), dict) else {}
    excluded = {str(item) for item in payload.get("excluded_models", [])}
    available = set(baselines.keys())
    if payload.get("proposed"):
        available.add("proposed")
    return [model for model in CANONICAL_MODEL_ORDER if model in available and model not in excluded]


def load_current_comparison_config(config_path: Path) -> dict[str, Any]:
    payload = yaml.safe_load(Path(config_path).read_text(encoding="utf-8")) or {}
    if not isinstance(payload, dict):
        raise ValueError(f"{config_path} 不包含有效 YAML object")
    return payload


def _ratio_override(settings: dict[str, Any], ratio_key: str = "8_2") -> dict[str, Any]:
    merged = dict(settings)
    overrides = settings.get("ratio_overrides", {})
    if isinstance(overrides, dict) and isinstance(overrides.get(ratio_key), dict):
        merged.update(overrides[ratio_key])
    merged.pop("ratio_overrides", None)
    return merged


def resolve_model_run_settings(config_payload: dict[str, Any], model_key: str, ratio_key: str = "8_2") -> dict[str, Any]:
    if model_key == "proposed":
        settings = config_payload.get("proposed", {})
    else:
        settings = (config_payload.get("baselines", {}) or {}).get(model_key, {})
    if not isinstance(settings, dict):
        raise KeyError(model_key)
    return _ratio_override(settings, ratio_key=ratio_key)


def load_reference_clean_rows_from_comparison_summaries(
    *,
    baseline_summary_path: Path,
    proposed_summary_path: Path,
    model_order: list[str],
    data_path: Path,
) -> dict[str, dict[str, Any]]:
    clean_rows: dict[str, dict[str, Any]] = {}
    baseline_rows = _read_csv_rows(baseline_summary_path)
    for row in baseline_rows:
        model_key = str(row.get("model", "")).strip()
        ratio = str(row.get("ratio", "")).strip()
        if model_key not in model_order or ratio != "8_2":
            continue
        metrics_path = str(row.get("metrics_path", "")).strip()
        result_dir = str(row.get("result_dir", "")).strip()
        checkpoint_path = ""
        if result_dir:
            result_dir_abs = _resolve_path_string(result_dir)
            if (Path(result_dir_abs) / "best.ckpt").exists():
                checkpoint_path = str((Path(result_dir_abs) / "best.ckpt").resolve())
        clean_rows[model_key] = {
            "ratio": "8:2",
            "model": model_key,
            "category": MODEL_CATEGORIES.get(model_key, ""),
            "snr_db": "clean",
            "snr_db_numeric": "",
            "actual_snr_db_mean": "",
            "noise_targets": "x_op+x_eis+x_cond",
            "test_accuracy": _pct_str_to_float(row.get("test_accuracy", row.get("accuracy", ""))),
            "accuracy_drop": 0.0,
            "test_macro_f1": _pct_str_to_float(row.get("test_macro_f1", row.get("macro_f1", ""))),
            "macro_f1_drop": 0.0,
            "test_weighted_f1": _pct_str_to_float(row.get("test_weighted_f1", row.get("weighted_f1", ""))),
            "weighted_f1_drop": 0.0,
            "test_inference_ms": _safe_float(row.get("test_inference_ms", row.get("inference_ms", 0.0))),
            "parameter_count": int(round(_safe_float(row.get("parameter_count", 0.0)))),
            "data_path": str(data_path),
            "metrics_path": metrics_path,
            "clean_alignment_source": "comparison_summary_reference",
            "result_dir": _resolve_path_string(result_dir) if result_dir else "",
            "checkpoint_path": checkpoint_path,
        }

    with proposed_summary_path.open("r", newline="", encoding="utf-8") as handle:
        for row in csv.DictReader(handle):
            ratio = str(row.get("ratio", "")).strip()
            if ratio != "8_2" or "proposed" not in model_order:
                continue
            output_dir_raw = str(row.get("output_dir", "")).strip()
            metrics_path = str(row.get("metrics_json", "")).strip()
            checkpoint_path = str(row.get("selected_ckpt", "")).strip()
            if metrics_path.upper() == "OK" and output_dir_raw:
                metrics_path = str(Path(output_dir_raw) / "metrics.json")
            if checkpoint_path.upper() == "OK" and output_dir_raw:
                checkpoint_path = str(Path(output_dir_raw) / "selected.ckpt")
            clean_rows["proposed"] = {
                "ratio": "8:2",
                "model": "proposed",
                "category": MODEL_CATEGORIES["proposed"],
                "snr_db": "clean",
                "snr_db_numeric": "",
                "actual_snr_db_mean": "",
                "noise_targets": "x_op+x_eis+x_cond",
                "test_accuracy": _pct_str_to_float(row.get("test_accuracy", "")),
                "accuracy_drop": 0.0,
                "test_macro_f1": _pct_str_to_float(row.get("test_macro_f1", "")),
                "macro_f1_drop": 0.0,
                "test_weighted_f1": _pct_str_to_float(row.get("test_weighted_f1", "")),
                "weighted_f1_drop": 0.0,
                "test_inference_ms": _safe_float(row.get("test_inference_ms", 0.0)),
                "parameter_count": int(round(_safe_float(row.get("parameter_count", 0.0)))),
                "data_path": str(data_path),
                "metrics_path": metrics_path,
                "clean_alignment_source": "comparison_summary_reference",
                "result_dir": _resolve_path_string(output_dir_raw),
                "checkpoint_path": _resolve_path_string(checkpoint_path),
            }
            break
    return clean_rows


def choose_clean_row_for_noise_alignment(
    reference_row: dict[str, Any],
    recomputed_row: dict[str, Any],
    tolerance: float = 1e-9,
) -> dict[str, Any]:
    metric_keys = ["test_accuracy", "test_macro_f1", "test_weighted_f1"]
    is_consistent = all(
        abs(_safe_float(reference_row.get(key)) - _safe_float(recomputed_row.get(key))) <= float(tolerance)
        for key in metric_keys
    )
    chosen = dict(reference_row if is_consistent else recomputed_row)
    chosen["clean_alignment_source"] = "comparison_summary_reference_validated" if is_consistent else "recomputed_due_to_mismatch"
    return chosen


def build_proposed_modality_summary_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    single_targets = {"x_op", "x_eis", "x_cond"}
    filtered: list[dict[str, Any]] = []
    for row in rows:
        if str(row.get("model", "")) != "proposed":
            continue
        if str(row.get("snr_db", "")).lower() == "clean":
            continue
        if str(row.get("noise_targets", "")) not in single_targets:
            continue
        filtered.append(dict(row))
    return filtered


def parse_model_path_overrides(items: list[str] | None) -> dict[str, Path]:
    overrides: dict[str, Path] = {}
    for item in items or []:
        model, sep, raw_path = str(item).partition("=")
        if not sep or not model or not raw_path:
            raise ValueError(f"模型路径覆盖格式错误: {item}，应为 model=path")
        overrides[model] = Path(raw_path)
    return overrides


def noise_seed_for_snr(seed: int, snr_db: float) -> int:
    return int(seed) * 1_000_000 + int(round(float(snr_db) * 1000.0))


def resolve_reuse_clean_artifacts(model_key: str, model_dir: Path) -> tuple[Path, Path]:
    metrics_path = Path(model_dir) / "metrics.json"
    checkpoint_path = Path(model_dir) / "best.ckpt"
    if not metrics_path.exists():
        raise FileNotFoundError(f"{model_key} 复用目录缺少 metrics.json: {metrics_path}")
    if not checkpoint_path.exists():
        raise FileNotFoundError(f"{model_key} 复用目录缺少 best.ckpt: {checkpoint_path}")
    return metrics_path, checkpoint_path


def load_reference_clean_rows(summary_path: Path, model_order: list[str], data_path: Path) -> dict[str, dict[str, Any]]:
    rows = _read_csv_rows(summary_path)
    clean_rows: dict[str, dict[str, Any]] = {}
    for row in rows:
        model_key = str(row.get("model", ""))
        if model_key not in model_order:
            continue
        metrics_path_raw = str(row.get("metrics_path", "")).strip()
        if not metrics_path_raw:
            continue
        metrics_path = Path(metrics_path_raw)
        if not metrics_path.exists():
            continue
        clean_rows[model_key] = _clean_row_from_metrics(model_key, metrics_path, data_path)
    return clean_rows


def _snr_token(snr_db: float) -> str:
    value = float(snr_db)
    if value.is_integer():
        return str(int(value))
    return f"{value:g}"


def _safe_float(value: Any, default: float = 0.0) -> float:
    try:
        if value in ("", None):
            return float(default)
        return float(value)
    except (TypeError, ValueError):
        return float(default)


def _pct(value: Any) -> str:
    return f"{_safe_float(value) * 100.0:.2f}%"


def _fmt_ms(value: Any) -> str:
    return f"{_safe_float(value):.3f}"


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", newline="", encoding="utf-8") as handle:
        return [normalize_summary_row_schema(row) for row in csv.DictReader(handle)]


def normalize_summary_row_schema(row: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(row)
    model_key = str(normalized.get("model", ""))
    if "test_accuracy" not in normalized and "accuracy" in normalized:
        normalized["test_accuracy"] = normalized.get("accuracy", "")
    if "test_macro_f1" not in normalized and "macro_f1" in normalized:
        normalized["test_macro_f1"] = normalized.get("macro_f1", "")
    if "test_weighted_f1" not in normalized and "weighted_f1" in normalized:
        normalized["test_weighted_f1"] = normalized.get("weighted_f1", "")
    if "test_inference_ms" not in normalized and "inference_ms" in normalized:
        normalized["test_inference_ms"] = normalized.get("inference_ms", "")
    if "category" not in normalized or not str(normalized.get("category", "")).strip():
        normalized["category"] = MODEL_CATEGORIES.get(model_key, "")
    if "ratio" not in normalized or not str(normalized.get("ratio", "")).strip():
        normalized["ratio"] = "8:2"
    if "noise_targets" not in normalized or not str(normalized.get("noise_targets", "")).strip():
        normalized["noise_targets"] = "x_op+x_eis+x_cond"
    return normalized


def _write_csv_rows(path: Path, rows: Iterable[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def write_clean_test_subset(clean_npz: Path, output_npz: Path) -> dict[str, Any]:
    with np.load(clean_npz) as data:
        if "split" not in data:
            raise ValueError(f"{clean_npz} 缺少 split 数组")
        split = np.asarray(data["split"], dtype=np.int64)
        test_idx = np.where(split == 2)[0]
        if test_idx.size == 0:
            raise ValueError(f"{clean_npz} 没有 split=2 的测试样本")
        n_samples = int(split.shape[0])
        payload: dict[str, np.ndarray[Any, Any]] = {}
        for key in data.files:
            array = np.asarray(data[key])
            if array.ndim > 0 and int(array.shape[0]) == n_samples:
                payload[key] = array[test_idx].copy()
            else:
                payload[key] = array.copy()
    payload["split"] = np.full(test_idx.shape, 2, dtype=np.int64)
    output_npz.parent.mkdir(parents=True, exist_ok=True)
    np.savez_compressed(output_npz, **payload)
    return {"source_npz": str(clean_npz), "output_npz": str(output_npz), "num_test_samples": int(test_idx.size)}


def _actual_snr_db(clean: np.ndarray[Any, Any], noisy: np.ndarray[Any, Any]) -> float:
    signal_power = float(np.mean(np.square(clean.astype(np.float64))))
    noise_power = float(np.mean(np.square((noisy - clean).astype(np.float64))))
    if signal_power <= 0 or noise_power <= 0:
        return float("inf")
    return float(10.0 * math.log10(signal_power / noise_power))


def write_test_subset_with_snr_noise(
    clean_npz: Path,
    output_npz: Path,
    snr_db: float,
    noise_targets: list[str],
    seed: int,
) -> dict[str, Any]:
    with np.load(clean_npz) as data:
        if "split" not in data:
            raise ValueError(f"{clean_npz} 缺少 split 数组")
        split = np.asarray(data["split"], dtype=np.int64)
        test_idx = np.where(split == 2)[0]
        if test_idx.size == 0:
            raise ValueError(f"{clean_npz} 没有 split=2 的测试样本")
        n_samples = int(split.shape[0])
        payload: dict[str, np.ndarray[Any, Any]] = {}
        for key in data.files:
            array = np.asarray(data[key])
            if array.ndim > 0 and int(array.shape[0]) == n_samples:
                payload[key] = array[test_idx].copy()
            else:
                payload[key] = array.copy()
    payload["split"] = np.full(test_idx.shape, 2, dtype=np.int64)

    rng = np.random.default_rng(int(seed))
    actual_by_target: dict[str, float] = {}
    for key in noise_targets:
        if key not in payload:
            raise KeyError(f"{clean_npz} 中不存在可加噪字段 {key}")
        clean_array = payload[key].astype(np.float32, copy=True)
        signal_power = float(np.mean(np.square(clean_array.astype(np.float64))))
        if signal_power <= 1e-12:
            signal_power = 1.0
        target_noise_power = signal_power / (10.0 ** (float(snr_db) / 10.0))
        noise = rng.normal(loc=0.0, scale=1.0, size=clean_array.shape).astype(np.float32)
        current_noise_power = float(np.mean(np.square(noise.astype(np.float64))))
        noise *= math.sqrt(target_noise_power / max(current_noise_power, 1e-12))
        noisy_array = clean_array + noise
        payload[key] = noisy_array.astype(np.float32)
        actual_by_target[key] = _actual_snr_db(clean_array, payload[key])

    output_npz.parent.mkdir(parents=True, exist_ok=True)
    np.savez_compressed(output_npz, **payload)
    return {
        "source_npz": str(clean_npz),
        "output_npz": str(output_npz),
        "snr_db": float(snr_db),
        "target_actual_snr_db": actual_by_target,
        "actual_snr_db_mean": float(np.mean(list(actual_by_target.values()))) if actual_by_target else float("nan"),
        "num_test_samples": int(test_idx.size),
    }


def _metric_payload(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError(f"{path} 不包含 JSON object")
    return payload


def _test_metrics(payload: dict[str, Any]) -> dict[str, Any]:
    test_payload = payload.get("test")
    if isinstance(test_payload, dict):
        return test_payload
    return payload


def _class0(metrics: dict[str, Any], key: str) -> float:
    for item in metrics.get("per_class_f1", []):
        if int(item.get("class_id", -1)) == 0:
            if key == "precision":
                return float(item.get("precision", 0.0))
            if key == "recall":
                return float(item.get("recall", 0.0))
            if key == "f1":
                return float(item.get("f1", 0.0))
    report = metrics.get("classification_report", {})
    class0 = report.get("0", {}) if isinstance(report, dict) else {}
    return float(class0.get({"precision": "precision", "recall": "recall", "f1": "f1-score"}[key], 0.0))


def _row_from_metrics(
    *,
    model_key: str,
    snr_db: float,
    actual_snr_db_mean: float,
    noise_targets: list[str],
    data_path: Path,
    metrics_path: Path,
    clean_row: dict[str, Any],
    clean_alignment_source: str,
    ratio_label: str = "8:2",
) -> dict[str, Any]:
    payload = _metric_payload(metrics_path)
    metrics = _test_metrics(payload)
    clean_acc = _safe_float(clean_row.get("test_accuracy"))
    clean_macro = _safe_float(clean_row.get("test_macro_f1"))
    clean_weighted = _safe_float(clean_row.get("test_weighted_f1"))
    test_acc = float(metrics.get("accuracy", 0.0))
    test_macro = float(metrics.get("macro_f1", 0.0))
    test_weighted = float(metrics.get("weighted_f1", 0.0))
    return {
        "ratio": ratio_label,
        "model": model_key,
        "category": MODEL_CATEGORIES[model_key],
        "snr_db": _snr_token(snr_db),
        "snr_db_numeric": float(snr_db),
        "actual_snr_db_mean": float(actual_snr_db_mean),
        "noise_targets": "+".join(noise_targets),
        "test_accuracy": test_acc,
        "accuracy_drop": clean_acc - test_acc,
        "test_macro_f1": test_macro,
        "macro_f1_drop": clean_macro - test_macro,
        "test_weighted_f1": test_weighted,
        "weighted_f1_drop": clean_weighted - test_weighted,
        "test_inference_ms": float(metrics.get("inference_time_per_sample_ms", 0.0)),
        "parameter_count": int(payload.get("parameter_count", clean_row.get("parameter_count", 0) or 0)),
        "data_path": str(data_path),
        "metrics_path": str(metrics_path),
        "clean_alignment_source": clean_alignment_source,
    }


def _clean_row_from_metrics(
    model_key: str,
    metrics_path: Path,
    data_path: Path,
    clean_alignment_source: str | None = None,
    ratio_label: str = "8:2",
) -> dict[str, Any]:
    payload = _metric_payload(metrics_path)
    metrics = _test_metrics(payload)
    return {
        "ratio": ratio_label,
        "model": model_key,
        "category": MODEL_CATEGORIES[model_key],
        "snr_db": "clean",
        "snr_db_numeric": "",
        "actual_snr_db_mean": "",
        "noise_targets": "x_op+x_eis+x_cond",
        "test_accuracy": float(metrics.get("accuracy", 0.0)),
        "accuracy_drop": 0.0,
        "test_macro_f1": float(metrics.get("macro_f1", 0.0)),
        "macro_f1_drop": 0.0,
        "test_weighted_f1": float(metrics.get("weighted_f1", 0.0)),
        "weighted_f1_drop": 0.0,
        "test_inference_ms": float(metrics.get("inference_time_per_sample_ms", 0.0)),
        "parameter_count": int(payload.get("parameter_count", 0)),
        "data_path": str(data_path),
        "metrics_path": str(metrics_path),
        "clean_alignment_source": str(clean_alignment_source or metrics_path),
    }


def _fmt4(value: Any) -> str:
    return f"{_safe_float(value):.4f}"


def export_summary_row(row: dict[str, Any]) -> dict[str, str]:
    return {
        "model": str(row.get("model", "")),
        "snr_db": str(row.get("snr_db", "")),
        "test_accuracy": _pct(row.get("test_accuracy")),
        "accuracy_drop": _pct(row.get("accuracy_drop")),
        "test_macro_f1": _pct(row.get("test_macro_f1")),
        "macro_f1_drop": _pct(row.get("macro_f1_drop")),
        "test_weighted_f1": _pct(row.get("test_weighted_f1")),
        "weighted_f1_drop": _pct(row.get("weighted_f1_drop")),
        "test_inference_ms": _fmt4(row.get("test_inference_ms")),
        "parameter_count": str(int(round(_safe_float(row.get("parameter_count", 0.0))))),
        "data_path": str(row.get("data_path", "")),
        "metrics_path": str(row.get("metrics_path", "")),
        "alignment_source": str(row.get("clean_alignment_source", row.get("alignment_source", ""))),
    }


def compact_summary_row(row: dict[str, Any]) -> dict[str, str]:
    return {
        "model": str(row.get("model", "")),
        "snr_db": str(row.get("snr_db", "")),
        "accuracy": _fmt4(row.get("test_accuracy")),
        "accuracy_drop": _fmt4(row.get("accuracy_drop")),
        "macro_f1": _fmt4(row.get("test_macro_f1")),
        "macro_f1_drop": _fmt4(row.get("macro_f1_drop")),
        "weighted_f1": _fmt4(row.get("test_weighted_f1")),
        "weighted_f1_drop": _fmt4(row.get("weighted_f1_drop")),
        "data_path": str(row.get("data_path", "")),
        "metrics_path": str(row.get("metrics_path", "")),
        "clean_alignment_source": str(row.get("clean_alignment_source", "")),
    }


def write_compact_summary(path: Path, rows: Iterable[dict[str, Any]]) -> None:
    _write_csv_rows(path, (compact_summary_row(row) for row in rows), COMPACT_SUMMARY_FIELDNAMES)


def write_export_summary(path: Path, rows: Iterable[dict[str, Any]]) -> None:
    _write_csv_rows(path, (export_summary_row(row) for row in rows), EXPORT_FIELDNAMES)


def merge_clean_rows_with_noise_rows(
    *,
    old_summary_path: Path | None = None,
    clean_rows_by_model: dict[str, dict[str, Any]] | None = None,
    noise_rows: list[dict[str, Any]],
    model_order: list[str],
    snr_order: list[str],
) -> list[dict[str, Any]]:
    if clean_rows_by_model is None:
        if old_summary_path is None:
            raise ValueError("需要提供 old_summary_path 或 clean_rows_by_model")
        old_rows = _read_csv_rows(old_summary_path)
        clean_rows = [row for row in old_rows if str(row.get("snr_db", "")).lower() == "clean" and row.get("model") in model_order]
        clean_rows_by_model = {str(row["model"]): dict(row) for row in clean_rows}
    by_key: dict[tuple[str, str], dict[str, Any]] = {
        (model, "clean"): dict(row) for model, row in clean_rows_by_model.items() if model in model_order
    }
    for row in noise_rows:
        by_key[(str(row["model"]), str(row["snr_db"]))] = row
    ordered: list[dict[str, Any]] = []
    for snr_label in snr_order:
        for model in model_order:
            row = by_key.get((model, snr_label))
            if row is not None:
                ordered.append(row)
    return ordered


def merge_updated_rows_with_existing_rows(
    *,
    existing_rows: list[dict[str, Any]],
    updated_rows: list[dict[str, Any]],
    model_order: list[str],
    snr_order: list[str],
) -> list[dict[str, Any]]:
    by_key: dict[tuple[str, str], dict[str, Any]] = {}
    for row in existing_rows:
        model = str(row.get("model", ""))
        snr_db = str(row.get("snr_db", ""))
        if model in model_order and snr_db in snr_order:
            by_key[(model, snr_db)] = dict(row)
    for row in updated_rows:
        by_key[(str(row.get("model", "")), str(row.get("snr_db", "")))] = dict(row)
    ordered: list[dict[str, Any]] = []
    for snr_label in snr_order:
        for model in model_order:
            row = by_key.get((model, snr_label))
            if row is not None:
                ordered.append(row)
    return ordered


def summary_row_to_workbook_values(row: dict[str, Any]) -> list[Any]:
    return [
        MODEL_DISPLAY_NAMES.get(str(row.get("model")), str(row.get("model"))),
        str(row.get("snr_db", "")),
        _pct(row.get("test_accuracy")),
        _pct(row.get("accuracy_drop")),
        _pct(row.get("test_macro_f1")),
        _pct(row.get("macro_f1_drop")),
        _pct(row.get("test_weighted_f1")),
        _fmt_ms(row.get("test_inference_ms")),
    ]


def update_workbook_snr_sheet(workbook_path: Path, rows: list[dict[str, Any]], sheet_name: str = "SNR噪声对比") -> None:
    workbook = load_workbook(workbook_path)
    try:
        if sheet_name not in workbook.sheetnames:
            worksheet = workbook.create_sheet(sheet_name)
            worksheet.append(["模型", "SNR(dB)", "Accuracy", "Acc下降", "Macro-F1", "Macro下降", "Weighted-F1", "推理时间(ms/sample)"])
        else:
            worksheet = workbook[sheet_name]
        if worksheet.max_row > 1:
            worksheet.delete_rows(2, worksheet.max_row - 1)
        for row in rows:
            worksheet.append(summary_row_to_workbook_values(row))
        for worksheet_row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in worksheet_row:
                font = copy(cell.font)
                font.name = "Times New Roman"
                cell.font = font
        workbook.save(workbook_path)
    finally:
        workbook.close()


def _copy_clean_rows_to_metrics(output_root: Path, clean_rows: list[dict[str, Any]]) -> None:
    for row in clean_rows:
        model = str(row.get("model"))
        source_raw = str(row.get("metrics_path", ""))
        source = Path(_resolve_path_string(source_raw)) if source_raw else Path()
        if not source.exists():
            continue
        clean_dir = output_root / model / "clean"
        clean_dir.mkdir(parents=True, exist_ok=True)
        target = clean_dir / "metrics.json"
        if source.resolve() != target.resolve():
            shutil.copy2(source, target)


def _normalize_feature_scope(model_key: str, settings: dict[str, Any] | None = None) -> str:
    payload = settings or {}
    raw_scope = str(payload.get("feature_scope", "")).strip().lower()
    if not raw_scope:
        return "x_op_only" if model_key == "logreg" else "all_modalities"
    aliases = {
        "x_op": "x_op_only",
        "x_op_only": "x_op_only",
        "op_only": "x_op_only",
        "all": "all_modalities",
        "all_modalities": "all_modalities",
        "x_op+x_eis+x_cond": "all_modalities",
        "multimodal": "all_modalities",
    }
    return aliases.get(raw_scope, raw_scope)


def _flatten_split_for_model(
    npz_path: Path,
    split_value: int,
    model_key: str,
    settings: dict[str, Any] | None = None,
) -> tuple[np.ndarray[Any, Any], np.ndarray[Any, Any]]:
    data = np.load(npz_path)
    split = np.asarray(data["split"], dtype=np.int64)
    indices = np.where(split == int(split_value))[0]
    cond_key = "x_cond" if "x_cond" in data else "cond"
    label_key = "labels" if "labels" in data else "y"
    x_op = np.asarray(data["x_op"][indices], dtype=np.float32)
    x_eis = np.asarray(data["x_eis"][indices], dtype=np.float32)
    x_cond = np.asarray(data[cond_key][indices], dtype=np.float32)
    feature_scope = _normalize_feature_scope(model_key, settings)
    if feature_scope == "x_op_only":
        features = x_op.reshape(len(indices), -1)
    elif feature_scope == "x_eis_only":
        features = x_eis.reshape(len(indices), -1)
    elif feature_scope == "x_cond_only":
        features = x_cond.reshape(len(indices), -1)
    elif feature_scope == "x_op+x_eis":
        features = np.concatenate(
            [
                x_op.reshape(len(indices), -1),
                x_eis.reshape(len(indices), -1),
            ],
            axis=1,
        )
    elif feature_scope == "x_op+x_cond":
        features = np.concatenate(
            [
                x_op.reshape(len(indices), -1),
                x_cond.reshape(len(indices), -1),
            ],
            axis=1,
        )
    elif feature_scope == "x_eis+x_cond":
        features = np.concatenate(
            [
                x_eis.reshape(len(indices), -1),
                x_cond.reshape(len(indices), -1),
            ],
            axis=1,
        )
    else:
        features = np.concatenate(
            [
                x_op.reshape(len(indices), -1),
                x_eis.reshape(len(indices), -1),
                x_cond.reshape(len(indices), -1),
            ],
            axis=1,
        )
    labels = np.asarray(data[label_key][indices], dtype=np.int64)
    return features, labels


def _normalize_class_weight(value: Any) -> str | None:
    raw = str(value or "").strip()
    if not raw or raw.lower() == "none":
        return None
    return raw


def _resolve_pca_components(value: Any) -> int | None:
    if value in (None, "", "None", "none", 0, "0", False):
        return None
    return int(value)


def _build_ml_model_from_settings(model_key: str, settings: dict[str, Any]) -> Any:
    use_scaler = bool(settings.get("use_scaler", model_key != "random_forest"))
    pca_components = _resolve_pca_components(settings.get("pca_components", 1))
    class_weight = _normalize_class_weight(settings.get("class_weighting", settings.get("class_weight", "none")))
    steps: list[tuple[str, Any]] = []
    if use_scaler:
        steps.append(("scaler", StandardScaler()))
    if pca_components is not None:
        steps.append(("pca", PCA(n_components=pca_components)))
    if model_key == "logreg":
        classifier = LogisticRegression(
            C=float(settings.get("C", 1.0)),
            max_iter=int(settings.get("max_iter", 1000)),
            class_weight=class_weight,
            random_state=int(settings.get("random_state", 44)),
        )
        steps.append(("classifier", classifier))
        return Pipeline(steps)
    if model_key == "svm":
        kernel = str(settings.get("kernel", "linear")).strip().lower()
        if kernel in {"linear", "linear_svc", "linearsvc"}:
            classifier = LinearSVC(
                C=float(settings.get("C", 1.0)),
                max_iter=int(settings.get("max_iter", 5000)),
                class_weight=class_weight,
                random_state=int(settings.get("random_state", 44)),
            )
        else:
            classifier = SVC(
                kernel=kernel,
                C=float(settings.get("C", 1.0)),
                gamma=settings.get("gamma", "scale"),
                class_weight=class_weight,
                random_state=int(settings.get("random_state", 44)),
            )
        steps.append(("classifier", classifier))
        return Pipeline(steps)
    if model_key == "random_forest":
        classifier = RandomForestClassifier(
            n_estimators=int(settings.get("n_estimators", 40)),
            max_depth=None if settings.get("max_depth") in (None, "", "None") else int(settings.get("max_depth")),
            min_samples_leaf=int(settings.get("min_samples_leaf", 1)),
            max_features=settings.get("max_features", "sqrt"),
            class_weight=class_weight,
            random_state=int(settings.get("random_state", 44)),
            n_jobs=-1,
        )
        if not steps:
            return classifier
        steps.append(("classifier", classifier))
        return Pipeline(steps)
    raise KeyError(model_key)


def _resolve_checkpoint_from_clean_row(clean_row: dict[str, Any]) -> Path | None:
    checkpoint_raw = str(clean_row.get("checkpoint_path", "")).strip()
    if checkpoint_raw:
        checkpoint_path = Path(_resolve_path_string(checkpoint_raw))
        if checkpoint_path.exists():
            return checkpoint_path
    result_dir_raw = str(clean_row.get("result_dir", "")).strip()
    if result_dir_raw:
        result_dir = Path(_resolve_path_string(result_dir_raw))
        for candidate_name in ("selected.ckpt", "best.ckpt", "best_val.ckpt"):
            candidate = result_dir / candidate_name
            if candidate.exists():
                return candidate
    metrics_path_raw = str(clean_row.get("metrics_path", "")).strip()
    if metrics_path_raw:
        metrics_parent = Path(_resolve_path_string(metrics_path_raw)).parent
        for candidate_name in ("selected.ckpt", "best.ckpt", "best_val.ckpt"):
            candidate = metrics_parent / candidate_name
            if candidate.exists():
                return candidate
    return None


def _train_and_evaluate(
    args: argparse.Namespace,
    noisy_npzs: dict[str, tuple[Path, dict[str, Any]]],
    clean_by_model: dict[str, dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]], list[dict[str, Any]]]:
    from scripts import run_official_baseline_experiments as baseline
    import torch
    from torch.utils.data import DataLoader

    import evaluate as proposed_eval

    output_root = ROOT / args.output_root
    clean_npz = ROOT / args.clean_npz
    clean_eval_npz = ROOT / args.data_root / "clean.npz"
    data_root = ROOT / args.data_root
    comparison_config = getattr(args, "_comparison_config", load_current_comparison_config(ROOT / args.comparison_config))
    ratio_key = str(args.ratio_key)
    ratio_label = str(args.ratio_label)
    noise_rows: list[dict[str, Any]] = []
    modality_rows: list[dict[str, Any]] = []
    fresh_clean_by_model: dict[str, dict[str, Any]] = dict(clean_by_model)
    ml_models = {"logreg", "svm", "random_forest"}
    torch_models = {"mlp", "cnn1d", "lstm", "transformer", "itransformer"}

    trained_models: dict[str, Any] = {}
    torch_checkpoints: dict[str, tuple[Path, Any]] = {}
    proposed_checkpoint: Path | None = None

    for model_key in args.models:
        print(f"\n=== 准备模型: {model_key} ===", flush=True)
        if model_key in ml_models:
            settings = resolve_model_run_settings(comparison_config, model_key, ratio_key=ratio_key)
            x_train, y_train = _flatten_split_for_model(clean_npz, split_value=0, model_key=model_key, settings=settings)
            model = _build_ml_model_from_settings(model_key, settings)
            model.fit(x_train, y_train)
            trained_models[model_key] = model
            x_val, y_val = _flatten_split_for_model(clean_npz, split_value=1, model_key=model_key, settings=settings)
            x_test, y_test = _flatten_split_for_model(clean_eval_npz, split_value=2, model_key=model_key, settings=settings)
            import time

            start = time.perf_counter()
            val_preds = model.predict(x_val)
            val_elapsed = time.perf_counter() - start
            start = time.perf_counter()
            test_preds = model.predict(x_test)
            test_elapsed = time.perf_counter() - start
            num_classes = baseline._num_classes(clean_npz)
            clean_dir = output_root / model_key / "clean"
            clean_metrics_path = baseline._save_result(
                clean_dir,
                baseline._classification_metrics(y_val, val_preds, val_elapsed, num_classes),
                baseline._classification_metrics(y_test, test_preds, test_elapsed, num_classes),
                baseline._ml_parameter_count(model),
                extra_payload={"clean_recomputed_for_snr_alignment": True},
            )
            recomputed_clean_row = _clean_row_from_metrics(
                model_key,
                clean_metrics_path,
                clean_eval_npz,
                clean_alignment_source=str(clean_metrics_path),
                ratio_label=ratio_label,
            )
            if model_key in clean_by_model:
                fresh_clean_by_model[model_key] = choose_clean_row_for_noise_alignment(
                    clean_by_model[model_key],
                    recomputed_clean_row,
                    tolerance=float(args.clean_consistency_tolerance),
                )
            else:
                fresh_clean_by_model[model_key] = recomputed_clean_row
        elif model_key in torch_models:
            settings = resolve_model_run_settings(comparison_config, model_key, ratio_key=ratio_key)
            train_ds = baseline.BaselineNPZDataset(clean_npz, split_value=0)
            num_classes = baseline._num_classes(clean_npz)
            checkpoint_path = _resolve_checkpoint_from_clean_row(clean_by_model.get(model_key, {}))
            if checkpoint_path is None:
                train_dir = output_root / "_trained_clean" / model_key
                metrics_path = baseline.run_torch_baseline(
                    model_key=model_key,
                    npz_path=clean_npz,
                    output_dir=train_dir,
                    epochs=int(settings.get("epochs", args.epochs)),
                    patience=int(settings.get("patience", args.patience)),
                    min_delta=float(args.min_delta),
                    batch_size=int(settings.get("batch_size", args.batch_size)),
                    lr=float(settings.get("lr", args.lr)),
                    weight_decay=float(settings.get("weight_decay", args.weight_decay)),
                    seed=int(args.seed),
                    hidden_dim=int(settings.get("hidden_dim", args.hidden_dim)),
                    d_model=int(settings.get("d_model", args.d_model)),
                    num_layers=int(settings.get("num_layers", args.num_layers)),
                    dropout=float(settings.get("dropout", args.dropout)),
                    refit_trainval=bool(args.refit_trainval),
                    min_epochs_before_stop=int(settings.get("min_epochs_before_stop", args.min_epochs_before_stop)),
                    val_metric_smoothing=int(settings.get("val_metric_smoothing", args.val_metric_smoothing)),
                    class_weighting=str(settings.get("class_weighting", args.class_weighting)),
                )
                checkpoint_path = output_root / "_trained_clean" / model_key / "best.ckpt"
            torch_checkpoints[model_key] = (checkpoint_path, (train_ds, num_classes, settings))
            if model_key not in clean_by_model:
                device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
                model = baseline._build_torch_model(
                    model_key,
                    train_ds,
                    num_classes,
                    int(settings.get("hidden_dim", args.hidden_dim)),
                    int(settings.get("d_model", args.d_model)),
                    int(settings.get("num_layers", args.num_layers)),
                    float(settings.get("dropout", args.dropout)),
                ).to(device)
                checkpoint = torch.load(checkpoint_path, map_location=device, weights_only=False)
                model.load_state_dict(checkpoint.get("model_state_dict", checkpoint), strict=True)
                val_ds = baseline.BaselineNPZDataset(clean_npz, split_value=1)
                test_ds = baseline.BaselineNPZDataset(clean_eval_npz, split_value=2)
                batch_size = int(settings.get("batch_size", args.batch_size))
                val_loader = DataLoader(val_ds, batch_size=batch_size, shuffle=False)
                test_loader = DataLoader(test_ds, batch_size=batch_size, shuffle=False)
                val_metrics = baseline._evaluate_torch_model(model, val_loader, device, num_classes)
                test_metrics = baseline._evaluate_torch_model(model, test_loader, device, num_classes)
                clean_metrics_path = baseline._save_result(
                    output_root / model_key / "clean",
                    val_metrics,
                    test_metrics,
                    int(sum(parameter.numel() for parameter in model.parameters())),
                    extra_payload={
                        "clean_recomputed_for_snr_alignment": True,
                        "clean_checkpoint_path": str(checkpoint_path),
                    },
                )
                fresh_clean_by_model[model_key] = _clean_row_from_metrics(
                    model_key,
                    clean_metrics_path,
                    clean_eval_npz,
                    clean_alignment_source=str(checkpoint_path),
                    ratio_label=ratio_label,
                )
        elif model_key == "proposed":
            settings = resolve_model_run_settings(comparison_config, model_key, ratio_key=ratio_key)
            proposed_checkpoint = _resolve_checkpoint_from_clean_row(clean_by_model.get(model_key, {}))
            if proposed_checkpoint is None:
                train_dir = output_root / "_trained_clean" / model_key
                baseline.run_proposed_model(
                    npz_path=clean_npz,
                    output_dir=train_dir,
                    config_path=ROOT / str(settings.get("config_file", args.proposed_config)),
                    epochs=int(settings.get("epochs", args.proposed_epochs)),
                    patience=int(settings.get("patience", args.patience)),
                    min_delta=float(args.min_delta),
                    batch_size=int(settings.get("batch_size", args.batch_size)),
                    refit_trainval=bool(args.refit_trainval),
                    min_epochs_before_stop=int(settings.get("min_epochs_before_stop", args.min_epochs_before_stop)),
                    val_metric_smoothing=int(settings.get("val_metric_smoothing", args.val_metric_smoothing)),
                    class_weighting=str(settings.get("class_weighting_default", args.class_weighting)),
                    seed=int(args.seed),
                    lr=float(settings.get("lr", args.proposed_lr)),
                    weight_decay=float(settings.get("weight_decay", args.proposed_weight_decay)),
                    checkpoint_selection=str(settings.get("checkpoint_selection", args.checkpoint_selection)),
                    selection_score=str(settings.get("selection_score_default", args.selection_score)),
                    init_checkpoint=args.init_checkpoint,
                )
                proposed_checkpoint = train_dir / "best.ckpt"
            if model_key not in clean_by_model:
                clean_dir = output_root / model_key / "clean"
                proposed_eval.run_evaluation(
                    config_path=str(ROOT / str(settings.get("config_file", args.proposed_config))),
                    data_path=str(clean_eval_npz),
                    checkpoint_path=str(proposed_checkpoint),
                    output_dir=str(clean_dir),
                    strict=False,
                    split="test",
                )
                clean_metrics_path = clean_dir / "metrics.json"
                fresh_clean_by_model[model_key] = _clean_row_from_metrics(
                    model_key,
                    clean_metrics_path,
                    clean_eval_npz,
                    clean_alignment_source=str(proposed_checkpoint),
                    ratio_label=ratio_label,
                )
        else:
            raise KeyError(model_key)

    for snr_label, (npz_path, noise_summary) in noisy_npzs.items():
        snr_db = float(noise_summary["snr_db"])
        print(f"\n=== 评估 SNR={snr_label} dB ===", flush=True)
        for model_key in args.models:
            clean_row = fresh_clean_by_model[model_key]
            eval_dir = output_root / model_key / f"snr_{snr_label}dB"
            eval_dir.mkdir(parents=True, exist_ok=True)
            if model_key in ml_models:
                model = trained_models[model_key]
                num_classes = baseline._num_classes(clean_npz)
                settings = resolve_model_run_settings(comparison_config, model_key, ratio_key=ratio_key)
                x_val, y_val = _flatten_split_for_model(clean_npz, split_value=1, model_key=model_key, settings=settings)
                x_test, y_test = _flatten_split_for_model(npz_path, split_value=2, model_key=model_key, settings=settings)
                import time

                start = time.perf_counter()
                val_preds = model.predict(x_val)
                val_elapsed = time.perf_counter() - start
                start = time.perf_counter()
                test_preds = model.predict(x_test)
                test_elapsed = time.perf_counter() - start
                val_metrics = baseline._classification_metrics(y_val, val_preds, val_elapsed, num_classes)
                test_metrics = baseline._classification_metrics(y_test, test_preds, test_elapsed, num_classes)
                metrics_path = baseline._save_result(
                    eval_dir,
                    val_metrics,
                    test_metrics,
                    baseline._ml_parameter_count(model),
                    extra_payload={"snr_noise_summary": noise_summary, "noise_evaluation": True},
                )
                clean_source = str(clean_row.get("metrics_path", ""))
            elif model_key in torch_models:
                checkpoint_path, (train_ds, num_classes, settings) = torch_checkpoints[model_key]
                device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
                model = baseline._build_torch_model(
                    model_key,
                    train_ds,
                    num_classes,
                    int(settings.get("hidden_dim", args.hidden_dim)),
                    int(settings.get("d_model", args.d_model)),
                    int(settings.get("num_layers", args.num_layers)),
                    float(settings.get("dropout", args.dropout)),
                ).to(device)
                checkpoint = torch.load(checkpoint_path, map_location=device, weights_only=False)
                model.load_state_dict(checkpoint.get("model_state_dict", checkpoint), strict=True)
                val_ds = baseline.BaselineNPZDataset(clean_npz, split_value=1)
                test_ds = baseline.BaselineNPZDataset(npz_path, split_value=2)
                batch_size = int(settings.get("batch_size", args.batch_size))
                val_loader = DataLoader(val_ds, batch_size=batch_size, shuffle=False)
                test_loader = DataLoader(test_ds, batch_size=batch_size, shuffle=False)
                val_metrics = baseline._evaluate_torch_model(model, val_loader, device, num_classes)
                test_metrics = baseline._evaluate_torch_model(model, test_loader, device, num_classes)
                param_count = int(sum(parameter.numel() for parameter in model.parameters()))
                metrics_path = baseline._save_result(
                    eval_dir,
                    val_metrics,
                    test_metrics,
                    param_count,
                    extra_payload={"snr_noise_summary": noise_summary, "noise_evaluation": True},
                )
                clean_source = str(checkpoint_path)
            else:
                if proposed_checkpoint is None:
                    raise RuntimeError("proposed checkpoint was not trained")
                settings = resolve_model_run_settings(comparison_config, model_key, ratio_key=ratio_key)
                proposed_eval.run_evaluation(
                    config_path=str(ROOT / str(settings.get("config_file", args.proposed_config))),
                    data_path=str(npz_path),
                    checkpoint_path=str(proposed_checkpoint),
                    output_dir=str(eval_dir),
                    strict=False,
                    split="test",
                )
                metrics_path = eval_dir / "metrics.json"
                clean_source = str(proposed_checkpoint)
            noise_rows.append(
                _row_from_metrics(
                    model_key=model_key,
                    snr_db=snr_db,
                    actual_snr_db_mean=float(noise_summary["actual_snr_db_mean"]),
                    noise_targets=list(args.noise_targets),
                    data_path=npz_path,
                    metrics_path=metrics_path,
                    clean_row=clean_row,
                    clean_alignment_source=clean_source,
                    ratio_label=ratio_label,
                )
            )
    if proposed_checkpoint is not None and "proposed" in fresh_clean_by_model:
        proposed_clean_row = fresh_clean_by_model["proposed"]
        proposed_settings = resolve_model_run_settings(comparison_config, "proposed", ratio_key=ratio_key)
        for modality_target in ["x_op", "x_eis", "x_cond"]:
            for snr_db in args.snr_dbs:
                label = _snr_token(snr_db)
                modality_npz = data_root / "proposed_modality" / modality_target / f"snr_{label}dB.npz"
                noise_summary = write_test_subset_with_snr_noise(
                    clean_npz,
                    modality_npz,
                    snr_db=float(snr_db),
                    noise_targets=[modality_target],
                    seed=noise_seed_for_snr(int(args.seed), float(snr_db)) + len(modality_target),
                )
                eval_dir = output_root / "proposed" / f"{modality_target}_snr_{label}dB"
                proposed_eval.run_evaluation(
                    config_path=str(ROOT / str(proposed_settings.get("config_file", args.proposed_config))),
                    data_path=str(modality_npz),
                    checkpoint_path=str(proposed_checkpoint),
                    output_dir=str(eval_dir),
                    strict=False,
                    split="test",
                )
                modality_rows.append(
                    _row_from_metrics(
                        model_key="proposed",
                        snr_db=float(snr_db),
                        actual_snr_db_mean=float(noise_summary["actual_snr_db_mean"]),
                        noise_targets=[modality_target],
                        data_path=modality_npz,
                        metrics_path=eval_dir / "metrics.json",
                        clean_row=proposed_clean_row,
                        clean_alignment_source=str(proposed_checkpoint),
                        ratio_label=ratio_label,
                    )
                )
    return noise_rows, fresh_clean_by_model, modality_rows


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="重新生成 SNR 噪声测试结果，并替换总表中的旧噪声结果")
    parser.add_argument("--clean-npz", default="data/processed/self_seed44_8_2.npz")
    parser.add_argument("--ratio-key", default="8_2", choices=["8_2", "7_3", "6_4", "5_5"])
    parser.add_argument("--ratio-label", default=None)
    parser.add_argument("--comparison-config", default="configs/current_comparison_models.yaml")
    parser.add_argument("--baseline-clean-summary", default="results/updated_dataset_baseline_ratio_comparison_20260513_seed44/test_summary.csv")
    parser.add_argument("--proposed-clean-summary", default="results/updated_dataset_proposed_ratio_comparison_20260513_seed44/proposed_summary.csv")
    parser.add_argument("--reference-clean-summary", default="results/updated_dataset_baseline_ratio_comparison_20260513_seed44/test_summary.csv")
    parser.add_argument("--old-summary", default="results/current_snr_noise_8_2_seed44_artifacts/summary.csv")
    parser.add_argument("--summary-path", default="results\\噪声对齐论文实验新表.csv")
    parser.add_argument("--output-root", default="results/current_snr_noise_8_2_seed44_artifacts")
    parser.add_argument("--data-root", default="data/processed/current_snr_noise_8_2_seed44_artifacts")
    parser.add_argument("--workbook", default="outputs/results_summary/CAPT-UniShape_实验结果总表_含自测公开训练结果汇总.xlsx")
    parser.add_argument("--clean-source", choices=["same-run", "reference"], default="reference", help="same-run：默认同次训练/同一checkpoint评估clean和噪声；reference：显式复用对比实验 clean 行")
    parser.add_argument("--models", nargs="+", default=DEFAULT_MODELS, choices=list(MODEL_CATEGORIES.keys()))
    parser.add_argument("--snr-dbs", nargs="+", type=float, default=DEFAULT_SNR_DBS)
    parser.add_argument("--noise-targets", nargs="+", default=["x_op", "x_eis", "x_cond"], choices=["x_op", "x_eis", "x_cond"])
    parser.add_argument("--seed", type=int, default=44)
    parser.add_argument("--epochs", type=int, default=80)
    parser.add_argument("--proposed-epochs", type=int, default=80)
    parser.add_argument("--patience", type=int, default=10)
    parser.add_argument("--min-delta", type=float, default=1e-4)
    parser.add_argument("--batch-size", type=int, default=32)
    parser.add_argument("--lr", type=float, default=1e-3)
    parser.add_argument("--weight-decay", type=float, default=1e-4)
    parser.add_argument("--proposed-lr", type=float, default=1e-4)
    parser.add_argument("--proposed-weight-decay", type=float, default=1e-4)
    parser.add_argument("--hidden-dim", type=int, default=64)
    parser.add_argument("--d-model", type=int, default=64)
    parser.add_argument("--num-layers", type=int, default=2)
    parser.add_argument("--dropout", type=float, default=0.1)
    parser.add_argument("--rf-estimators", type=int, default=200)
    parser.add_argument("--refit-trainval", action=argparse.BooleanOptionalAction, default=False)
    parser.add_argument("--min-epochs-before-stop", type=int, default=20)
    parser.add_argument("--val-metric-smoothing", type=int, default=3)
    parser.add_argument("--class-weighting", default="sqrt_balanced")
    parser.add_argument("--proposed-config", default="configs/rbf_kanfusion_optimized.yaml")
    parser.add_argument("--checkpoint-selection", choices=["best_val", "best-val", "last"], default="best_val")
    parser.add_argument("--selection-score", choices=["macro_f1", "macro_f1_class0_recall", "macro_f1_gap_penalty", "val_train_holdout_macro_f1"], default="macro_f1")
    parser.add_argument("--init-checkpoint", default=None)
    parser.add_argument("--reuse-torch-clean-dir", nargs="+", default=[], help="复用已有深度学习 clean 模型目录，格式 model=dir；目录下需包含 metrics.json 和 best.ckpt")
    parser.add_argument("--clean-consistency-tolerance", type=float, default=1e-9)
    parser.add_argument("--write-paper-table", action="store_true", help="额外输出 paper_snr_noise_table.md；默认只写 summary.csv")
    parser.add_argument("--write-proposed-modality-summary", action=argparse.BooleanOptionalAction, default=True, help="输出 proposed 单模态噪声副表")
    parser.add_argument("--clear-previous-results", action=argparse.BooleanOptionalAction, default=True, help="运行前清理旧的噪声实验中间产物和主表")
    parser.add_argument("--skip-workbook", action="store_true")
    args = parser.parse_args(argv)
    if args.ratio_label is None:
        args.ratio_label = str(args.ratio_key).replace("_", ":")
    return args


def main(argv: list[str] | None = None) -> None:
    args = parse_args(argv)
    output_root = ROOT / args.output_root
    data_root = ROOT / args.data_root
    clean_npz = ROOT / args.clean_npz
    old_summary_path = ROOT / args.old_summary
    summary_path = ROOT / args.summary_path
    compact_summary_path = output_root / "summary.csv"
    modality_summary_path = output_root / "proposed_modality_summary.csv"
    args._comparison_config = load_current_comparison_config(ROOT / args.comparison_config)
    available_models = load_model_order_from_current_comparison_config(ROOT / args.comparison_config)
    args.models = [model for model in args.models if model in available_models]
    if bool(args.clear_previous_results):
        if summary_path.exists():
            summary_path.unlink()
        if output_root.exists():
            shutil.rmtree(output_root)
        if data_root.exists():
            shutil.rmtree(data_root)
    output_root.mkdir(parents=True, exist_ok=True)
    data_root.mkdir(parents=True, exist_ok=True)

    old_rows = _read_csv_rows(old_summary_path) if old_summary_path.exists() else []
    clean_subset_summary = write_clean_test_subset(clean_npz, data_root / "clean.npz")
    clean_eval_npz = data_root / "clean.npz"
    clean_by_model: dict[str, dict[str, Any]] = {}
    if args.clean_source == "reference":
        clean_by_model = load_reference_clean_rows_from_comparison_summaries(
            baseline_summary_path=ROOT / args.baseline_clean_summary,
            proposed_summary_path=ROOT / args.proposed_clean_summary,
            model_order=list(args.models),
            data_path=clean_eval_npz,
        )
        _copy_clean_rows_to_metrics(output_root, list(clean_by_model.values()))

    noisy_npzs: dict[str, tuple[Path, dict[str, Any]]] = {}
    protocol = {
        "clean_npz": str(clean_npz),
        "old_summary": str(old_summary_path),
        "ratio": "8:2",
        "seed": int(args.seed),
        "snr_dbs": ["clean"] + [_snr_token(value) for value in args.snr_dbs],
        "noise_targets": list(args.noise_targets),
        "noise_scope": "test_split_only",
        "clean_rows_source": "same_run_same_checkpoint" if args.clean_source == "same-run" else "reference_clean_summary_explicit",
        "original_no_noise_workbook_sheets": "untouched",
        "reference_clean_summary": str(ROOT / args.reference_clean_summary),
        "clean_subset_summary": clean_subset_summary,
    }
    for snr_db in args.snr_dbs:
        label = _snr_token(snr_db)
        noisy_path = data_root / f"snr_{label}dB.npz"
        noise_summary = write_test_subset_with_snr_noise(
            clean_npz,
            noisy_path,
            snr_db=float(snr_db),
            noise_targets=list(args.noise_targets),
            seed=noise_seed_for_snr(int(args.seed), float(snr_db)),
        )
        noisy_npzs[label] = (noisy_path, noise_summary)
    (output_root / "protocol.json").write_text(json.dumps(protocol, indent=2, ensure_ascii=False), encoding="utf-8")

    noise_rows, fresh_clean_by_model, modality_rows = _train_and_evaluate(args, noisy_npzs, clean_by_model)
    missing_clean = [model for model in args.models if model not in fresh_clean_by_model]
    if missing_clean:
        raise ValueError(f"本次 SNR 汇总缺少同次 clean 行: {missing_clean}")
    updated_rows = merge_clean_rows_with_noise_rows(
        clean_rows_by_model=fresh_clean_by_model,
        noise_rows=noise_rows,
        model_order=list(args.models),
        snr_order=["clean"] + [_snr_token(value) for value in args.snr_dbs],
    )
    write_export_summary(summary_path, updated_rows)
    write_compact_summary(compact_summary_path, updated_rows)
    paper_path = output_root / "paper_snr_noise_table.md"
    if bool(args.write_proposed_modality_summary):
        _write_csv_rows(modality_summary_path, build_proposed_modality_summary_rows(modality_rows), SUMMARY_FIELDNAMES)
    if bool(args.write_paper_table):
        write_paper_table(paper_path, updated_rows, summary_path)
    if not bool(args.skip_workbook):
        update_workbook_snr_sheet(ROOT / args.workbook, updated_rows)
    print(f"\n已写入 SNR 噪声汇总: {summary_path}", flush=True)
    print(f"已写入紧凑汇总: {compact_summary_path}", flush=True)
    if bool(args.write_proposed_modality_summary):
        print(f"已写入 proposed 单模态副表: {modality_summary_path}", flush=True)
    if bool(args.write_paper_table):
        print(f"已更新论文表: {paper_path}", flush=True)
    if not bool(args.skip_workbook):
        print(f"已更新总表: {ROOT / args.workbook}", flush=True)


def write_paper_table(path: Path, rows: list[dict[str, Any]], summary_path: Path) -> None:
    lines = [
        "# SNR 噪声鲁棒性基线对比",
        "",
        "协议：clean 行直接引用 8:2 对比实验结论；噪声行仅对测试集 `x_op+x_eis+x_cond` 按目标 SNR 加高斯噪声后重新测试。",
        "",
        "| 模型 | SNR(dB) | Accuracy | Acc下降 | Macro-F1 | Macro下降 | Weighted-F1 | 推理时间(ms/sample) |",
        "|---|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for row in rows:
        values = summary_row_to_workbook_values(row)
        lines.append("| " + " | ".join(str(value) for value in values) + " |")
    lines.extend(["", f"数据来源：`{summary_path.as_posix()}`", ""])
    path.write_text("\n".join(lines), encoding="utf-8")


if __name__ == "__main__":
    main()
