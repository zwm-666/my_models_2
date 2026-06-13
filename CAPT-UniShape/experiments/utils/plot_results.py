"""Plot the official baseline comparison summary and related experiment figures."""

from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path
from typing import Any

import matplotlib.pyplot as plt
import numpy as np


ROOT = Path(__file__).resolve().parents[2]
FIG_STYLE_DIR = ROOT / "outputs" / "paper_figures"
if str(FIG_STYLE_DIR) not in sys.path:
    sys.path.insert(0, str(FIG_STYLE_DIR))

from plot_style import (  # noqa: E402
    ANNOTATION_SIZE,
    LEGEND_KWARGS,
    MODEL_COLORS,
    STACKED_SUBPLOT_ONE_LINE_HSPACE,
    add_bottom_caption,
    add_panel_tag,
    apply_paper_style,
    figsize_full,
    inside_text_y,
    save_paper_figure,
    show_shared_x_axis,
    style_axes,
    style_legend_frame,
    ylim_with_inside_text,
)


CLASS_NAMES = ["正常", "过干/膜干", "过湿/水淹"]
DEFAULT_RESULTS_WORKBOOK = Path("outputs/results_summary/实验结果总表.xlsx")
PAPER_OUTPUT_FORMATS = ("png", "svg")

METRIC_SPECS = [
    ("(a)", "test_accuracy", "Accuracy (%)", 40.0, 100.0),
    ("(b)", "test_macro_f1", "Macro-F1 (%)", 10.0, 100.0),
    ("(c)", "test_weighted_f1", "Weighted-F1 (%)", 30.0, 100.0),
]


MODEL_ALIASES = {
    "proposed": "proposed",
    "capt-unishape": "proposed",
    "capt_unishape": "proposed",
    "xgboost": "xgboost",
    "lightgbm": "lightgbm",
    "logistic regression": "logreg",
    "logreg": "logreg",
    "lr": "logreg",
    "svm": "svm",
    "random forest": "random_forest",
    "random_forest": "random_forest",
    "rf": "random_forest",
    "mlp": "mlp",
    "1d-cnn": "cnn1d",
    "cnn1d": "cnn1d",
    "cnn": "cnn1d",
    "tcn": "tcn",
    "autoformer": "autoformer",
    "transformer": "transformer",
    "itransformer": "itransformer",
}

MODEL_ORDER = [
    "proposed",
    "xgboost",
    "lightgbm",
    "mlp",
    "tcn",
    "autoformer",
    "transformer",
    "itransformer",
    "logreg",
    "svm",
    "random_forest",
    "cnn1d",
]

MODEL_LABELS = {
    "proposed": "CAPT-UniShape",
    "xgboost": "XGBoost",
    "lightgbm": "LightGBM",
    "logreg": "LR",
    "svm": "SVM",
    "random_forest": "RF",
    "mlp": "MLP",
    "cnn1d": "CNN",
    "tcn": "TCN",
    "autoformer": "Autoformer",
    "transformer": "Transformer",
    "itransformer": "iTransformer",
}

MODEL_MARKERS = {
    "proposed": "o",
    "xgboost": "o",
    "lightgbm": "s",
    "mlp": "D",
    "tcn": "^",
    "autoformer": "v",
    "transformer": "P",
    "itransformer": "X",
    "logreg": "h",
    "svm": "p",
    "random_forest": "*",
    "cnn1d": "<",
}

MODEL_COLOR_MAP = {model: MODEL_COLORS[index % len(MODEL_COLORS)] for index, model in enumerate(MODEL_ORDER)}


def _read_summary_csv(summary_path: Path) -> list[dict[str, str]]:
    if not summary_path.is_file():
        raise FileNotFoundError(f"Missing summary CSV: {summary_path}")
    with summary_path.open("r", encoding="utf-8-sig", newline="") as file:
        return list(csv.DictReader(file))


def _safe_float(row: dict[str, str], key: str) -> float:
    value = row.get(key, "")
    return _coerce_float(value)


def _safe_metric_float(row: dict[str, str], key: str) -> float:
    fallback_keys = {
        "test_accuracy": "accuracy",
        "test_macro_f1": "macro_f1",
        "test_weighted_f1": "weighted_f1",
        "test_inference_ms": "inference_ms",
    }
    value = row.get(key, "")
    if value not in (None, ""):
        return _coerce_float(value)
    fallback_key = fallback_keys.get(key)
    if fallback_key is not None:
        return _coerce_float(row.get(fallback_key, ""))
    return _coerce_float(value)


def _fill_metric_aliases(row: dict[str, str]) -> dict[str, str]:
    normalized = dict(row)
    for test_key, compact_key in (
        ("test_accuracy", "accuracy"),
        ("test_macro_f1", "macro_f1"),
        ("test_weighted_f1", "weighted_f1"),
        ("test_inference_ms", "inference_ms"),
    ):
        test_value = normalized.get(test_key, "")
        compact_value = normalized.get(compact_key, "")
        if compact_value in (None, "") and test_value not in (None, ""):
            normalized[compact_key] = test_value
    return normalized


def _noise_axis_limits_and_ticks(
    metric_values: list[float],
    proposed_values: list[float],
    nominal_y_max: float,
) -> tuple[float, float, np.ndarray[Any, Any]]:
    finite_values = [float(value) for value in metric_values if np.isfinite(value)]
    if finite_values:
        y_min = max(0.0, float(np.floor((min(finite_values) - 2.0) / 5.0) * 5.0))
    else:
        y_min = 0.0
    proposed_top = max(proposed_values or [nominal_y_max])
    _, y_max = ylim_with_inside_text(proposed_top, y_min, upper=nominal_y_max)
    tick_values = np.arange(y_min, nominal_y_max + 0.1, 10.0)
    if tick_values.size == 0 or abs(float(tick_values[0]) - y_min) > 1e-9:
        tick_values = np.insert(tick_values, 0, y_min)
    if abs(float(tick_values[-1]) - float(nominal_y_max)) > 1e-9:
        tick_values = np.append(tick_values, float(nominal_y_max))
    return y_min, y_max, tick_values


def _coerce_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return 0.0
    if text.endswith("%"):
        return float(text[:-1]) / 100.0
    return float(text)


def _normalize_model_name(value: Any) -> str:
    text = str(value or "").strip()
    key = text.lower().replace(" ", "_")
    return MODEL_ALIASES.get(text.lower(), MODEL_ALIASES.get(key, key))


def _read_workbook_sheet_rows(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    if not workbook_path.is_file():
        return []
    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return []
        sheet = workbook[sheet_name]
        values = list(sheet.iter_rows(values_only=True))
        header_index = next(
            (
                index
                for index, row in enumerate(values)
                if row and any(cell is not None and str(cell).strip() for cell in row)
            ),
            None,
        )
        if header_index is None:
            return []
        headers = [str(cell).strip() if cell is not None else "" for cell in values[header_index]]
        rows: list[dict[str, Any]] = []
        for row in values[header_index + 1 :]:
            record = {headers[index]: row[index] if index < len(row) else None for index in range(len(headers)) if headers[index]}
            if any(value not in (None, "") for value in record.values()):
                rows.append(record)
        return rows
    finally:
        workbook.close()


def _workbook_baseline_rows(workbook_path: Path) -> list[dict[str, str]]:
    output: list[dict[str, str]] = []
    for row in _read_workbook_sheet_rows(workbook_path, "对比实验"):
        model = _normalize_model_name(row.get("模型名称"))
        if model == "proposed" or not model:
            continue
        ratio = str(row.get("比例") or "").strip().replace("_", ":")
        if not ratio:
            continue
        output.append(
            {
                "ratio": ratio,
                "model": model,
                "test_accuracy": f"{_coerce_float(row.get('Test Acc')):.10f}",
                "test_macro_f1": f"{_coerce_float(row.get('Macro-F1')):.10f}",
                "test_weighted_f1": f"{_coerce_float(row.get('Weighted-F1')):.10f}",
                "test_inference_ms": f"{_coerce_float(row.get('时间(ms/sample)')):.10f}",
                "parameter_count": str(int(_coerce_float(row.get("参数量")))),
                "metrics_path": str(row.get("结果来源") or ""),
            }
        )
    return output


def _workbook_noise_rows(workbook_path: Path) -> list[dict[str, str]]:
    output: list[dict[str, str]] = []
    for row in _read_workbook_sheet_rows(workbook_path, "SNR噪声对比"):
        model = _normalize_model_name(row.get("model"))
        if not model:
            continue
        snr_db = str(row.get("snr_db") or "").strip()
        if not snr_db:
            continue
        output.append(
            {
                "variant": model,
                "description": "",
                "scenario": "clean" if snr_db.lower() == "clean" else "workbook_noise",
                "snr_db": snr_db,
                "test_accuracy": f"{_coerce_float(row.get('test_accuracy')):.10f}",
                "test_macro_f1": f"{_coerce_float(row.get('test_macro_f1')):.10f}",
                "test_weighted_f1": f"{_coerce_float(row.get('test_weighted_f1')):.10f}",
                "test_inference_ms": f"{_coerce_float(row.get('test_inference_ms')):.10f}",
            }
        )
    return output


def _normalize_snr_proposed_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    output: list[dict[str, str]] = []
    for row in rows:
        variant = str(row.get("variant", "")).strip()
        normalized = _fill_metric_aliases(row)
        if variant == "full_rbf":
            normalized["variant"] = "proposed"
        output.append(normalized)
    return output


def _merge_metric_rows(
    primary_rows: list[dict[str, str]],
    extra_rows: list[dict[str, str]],
    key_fields: tuple[str, ...],
) -> list[dict[str, str]]:
    merged: dict[tuple[str, ...], dict[str, str]] = {}
    order: list[tuple[str, ...]] = []
    for row in extra_rows + primary_rows:
        key = tuple(str(row.get(field, "")) for field in key_fields)
        if key not in merged:
            order.append(key)
            merged[key] = dict(row)
            continue
        merged_row = dict(merged[key])
        for field, value in row.items():
            if value in (None, "") and field in merged_row:
                continue
            if field not in merged_row or merged_row.get(field, "") in (None, ""):
                merged_row[field] = value
            else:
                merged_row[field] = value
        merged[key] = merged_row
    return [merged[key] for key in order]


def _write_source_csv(path: Path, rows: list[dict[str, str]]) -> None:
    if not rows:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames: list[str] = []
    for row in rows:
        for key in row:
            if key not in fieldnames:
                fieldnames.append(key)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _load_metrics(results_dir: Path) -> dict[str, Any]:
    metrics_path = results_dir / "metrics.json"
    if not metrics_path.is_file():
        raise FileNotFoundError(f"Missing metrics file: {metrics_path}")
    return json.loads(metrics_path.read_text(encoding="utf-8"))


def _confusion_from_metrics(metrics: dict[str, Any]) -> np.ndarray[Any, Any]:
    if "test" in metrics and isinstance(metrics["test"], dict):
        return np.asarray(metrics["test"]["confusion_matrix"], dtype=np.int64)
    return np.asarray(metrics["confusion_matrix"], dtype=np.int64)


def _score_from_metrics(metrics: dict[str, Any], key: str) -> float:
    if "test" in metrics and isinstance(metrics["test"], dict) and key in metrics["test"]:
        return float(metrics["test"][key])
    return float(metrics[key])


def plot_confusion_matrix(results_dir: Path, output_dir: Path, title: str | None = None) -> Path:
    metrics = _load_metrics(results_dir)
    cm = _confusion_from_metrics(metrics)
    row_sum = np.maximum(cm.sum(axis=1, keepdims=True), 1)
    cm_norm = cm / row_sum
    labels = CLASS_NAMES[: cm.shape[0]]

    fig, ax = plt.subplots(figsize=(6.4, 5.4))
    image = ax.imshow(cm_norm, cmap="Blues", vmin=0.0, vmax=1.0)
    fig.colorbar(image, ax=ax, fraction=0.046, pad=0.04)
    for row in range(cm.shape[0]):
        for col in range(cm.shape[1]):
            ax.text(col, row, f"{cm[row, col]}\n{cm_norm[row, col] * 100:.1f}%", ha="center", va="center")
    ax.set_xticks(range(len(labels)))
    ax.set_yticks(range(len(labels)))
    ax.set_xticklabels(labels)
    ax.set_yticklabels(labels)
    ax.set_xlabel("预测类别")
    ax.set_ylabel("真实类别")
    ax.set_title(title or f"混淆矩阵：{results_dir.name}")
    fig.tight_layout()
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / f"{results_dir.name}_confusion_matrix.png"
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    return path


def plot_model_comparison(results_dirs: list[Path], output_dir: Path) -> Path:
    names: list[str] = []
    acc: list[float] = []
    macro_f1: list[float] = []
    params: list[float] = []
    for results_dir in results_dirs:
        metrics = _load_metrics(results_dir)
        names.append(results_dir.name.replace("official_unishape_", ""))
        acc.append(_score_from_metrics(metrics, "accuracy"))
        macro_f1.append(_score_from_metrics(metrics, "macro_f1"))
        params.append(float(metrics.get("parameter_count", 0)) / 1_000_000.0)

    x = np.arange(len(names))
    width = 0.35
    fig, ax = plt.subplots(figsize=(max(7.0, len(names) * 2.4), 5.0))
    bars_acc = ax.bar(x - width / 2, acc, width, label="准确率", color="#4e79a7")
    bars_f1 = ax.bar(x + width / 2, macro_f1, width, label="宏平均F1", color="#e15759")
    for bars in (bars_acc, bars_f1):
        for bar in bars:
            value = bar.get_height()
            ax.text(bar.get_x() + bar.get_width() / 2, value + 0.015, f"{value:.3f}", ha="center", va="bottom", fontsize=9)
    ax.set_xticks(x)
    ax.set_xticklabels(names, rotation=15, ha="right")
    ax.set_ylim(0, 1.05)
    ax.set_ylabel("得分")
    ax.set_title("官方 UniShape 版 CAPT-UniShape 模型对比")
    ax.legend()
    ax2 = ax.twinx()
    ax2.plot(x, params, marker="o", color="#59a14f", label="参数量（百万）")
    ax2.set_ylabel("参数量（百万）")
    fig.tight_layout()
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / "official_model_comparison.png"
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    return path


def plot_split_summary(
    summary_path: Path,
    output_dir: Path,
    title: str = "Model comparison under different training ratios",
    output_name: str = "official_split_ratio_summary.png",
    results_workbook: Path | None = DEFAULT_RESULTS_WORKBOOK,
) -> Path:
    apply_paper_style()
    rows = _read_summary_csv(summary_path)
    if results_workbook is not None:
        rows = _merge_metric_rows(rows, _workbook_baseline_rows(results_workbook), ("ratio", "model"))
    ratio_order = {"5_5": 0, "6_4": 1, "7_3": 2, "8_2": 3}
    ratio_labels = {"8_2": "80%", "7_3": "70%", "6_4": "60%", "5_5": "50%"}
    ratios = sorted({row["ratio"] for row in rows}, key=lambda item: ratio_order.get(item.replace(":", "_"), 99))
    reference_ratio = "5:5"
    reference_values = {
        row["model"]: _safe_float(row, "test_macro_f1")
        for row in rows
        if row.get("ratio", "").replace("_", ":") == reference_ratio
    }
    models = sorted(
        {row["model"] for row in rows},
        key=lambda item: (
            reference_values.get(item, float("inf")),
            MODEL_ORDER.index(item) if item in MODEL_ORDER else 99,
            item,
        ),
    )
    x = np.arange(len(ratios))

    fig, axes = plt.subplots(
        3,
        1,
        figsize=figsize_full(height_cm=16.0),
        sharex=True,
        gridspec_kw={"hspace": STACKED_SUBPLOT_ONE_LINE_HSPACE},
    )
    for ax, (panel_tag, metric_col, ylabel, y_floor, nominal_y_max) in zip(axes, METRIC_SPECS):
        metric_values = [
            _safe_float(row, metric_col) * 100.0
            for row in rows
            if row.get("model") in models and row.get("ratio") in ratios
        ]
        y_min = min(y_floor, max(0.0, float(np.floor((min(metric_values) - 2.0) / 10.0) * 10.0))) if metric_values else y_floor
        proposed_values: list[float] = []
        for ratio in ratios:
            matched = [row for row in rows if row["ratio"] == ratio and row["model"] == "proposed"]
            if matched:
                proposed_values.append(_safe_float(matched[-1], metric_col) * 100.0)
        _, y_max = ylim_with_inside_text(max(proposed_values or [nominal_y_max]), y_min, upper=nominal_y_max)
        for model in models:
            values = []
            for ratio in ratios:
                matched = [row for row in rows if row["ratio"] == ratio and row["model"] == model]
                values.append(_safe_float(matched[-1], metric_col) * 100.0 if matched else np.nan)
            is_proposed = model == "proposed"
            ax.plot(
                x,
                values,
                color=MODEL_COLOR_MAP.get(model, "black"),
                marker=MODEL_MARKERS.get(model, "o"),
                markersize=4.8 if is_proposed else 3.8,
                linewidth=2.0 if is_proposed else 1.05,
                markerfacecolor=MODEL_COLOR_MAP.get(model, "black") if is_proposed else "white",
                markeredgecolor=MODEL_COLOR_MAP.get(model, "black"),
                markeredgewidth=0.9,
                alpha=1.0 if is_proposed else 0.88,
                label=MODEL_LABELS.get(model, model),
                zorder=5 if is_proposed else 3,
            )
            if is_proposed:
                for xi, yi in zip(x, values):
                    if np.isnan(yi):
                        continue
                    ax.text(
                        xi,
                        inside_text_y(float(yi), y_min, y_max),
                        f"{yi:.2f}",
                        ha="center",
                        va="bottom",
                        fontsize=ANNOTATION_SIZE,
                        fontweight="bold",
                        color="black",
                        clip_on=True,
                    )
        ax.set_ylabel(ylabel)
        ax.set_ylim(y_min, y_max)
        ax.set_yticks(np.arange(int(y_min), int(nominal_y_max) + 1, 10))
        style_axes(ax, grid_axis="y")
        add_panel_tag(ax, panel_tag)
        show_shared_x_axis(ax, x, [ratio_labels.get(r.replace(":", "_"), r) for r in ratios], "Training set ratio")

    handles, labels = axes[0].get_legend_handles_labels()
    legend = axes[0].legend(
        handles,
        labels,
        ncol=4,
        loc="lower left",
        bbox_to_anchor=(0.0, 1.18, 1.0, 0.06),
        bbox_transform=axes[0].transAxes,
        mode="expand",
        handlelength=1.6,
        columnspacing=1.0,
        borderaxespad=0,
        **LEGEND_KWARGS,
    )
    style_legend_frame(legend)
    add_bottom_caption(fig, "(a) Accuracy; (b) Macro-F1; (c) Weighted-F1.")
    fig.subplots_adjust(left=0.10, right=0.98, top=0.84, bottom=0.16, hspace=STACKED_SUBPLOT_ONE_LINE_HSPACE)
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / output_name
    save_paper_figure(fig, str(path.with_suffix("")), formats=PAPER_OUTPUT_FORMATS)
    plt.close(fig)
    return path


def plot_baseline_summary(summary_path: Path, output_dir: Path, results_workbook: Path | None = DEFAULT_RESULTS_WORKBOOK) -> Path:
    return plot_split_summary(
        summary_path,
        output_dir,
        title="",
        output_name="official_baseline_comparison_summary.png",
        results_workbook=results_workbook,
    )


def plot_noise_summary(summary_path: Path, output_dir: Path, results_workbook: Path | None = DEFAULT_RESULTS_WORKBOOK) -> Path:
    apply_paper_style()
    rows = _read_summary_csv(summary_path)
    if rows and "variant" in rows[0] and "snr_db" in rows[0]:
        rows = _normalize_snr_proposed_rows(rows)
        if results_workbook is not None:
            comparison_rows = [row for row in _workbook_noise_rows(results_workbook) if row.get("variant") != "proposed"]
            rows = _merge_metric_rows(rows, comparison_rows, ("variant", "snr_db"))
        model_key = "variant"
        x_key = "snr_db"
        x_label = "SNR (dB)"
        x_values = sorted(
            {row[x_key] for row in rows},
            key=lambda value: float("inf") if str(value).lower() == "clean" else float(value),
        )
        x_positions = np.arange(len(x_values))
        x_ticklabels = [str(value) for value in x_values]
        snr_order = {str(value): index for index, value in enumerate(x_values)}
        rows = sorted(rows, key=lambda row: (snr_order.get(str(row.get(x_key, "")), 999), str(row.get(model_key, ""))))
    else:
        model_key = "model"
        x_key = "noise_std"
        x_label = "Relative Gaussian noise standard deviation"
        numeric_values = sorted({_safe_float(row, x_key) for row in rows})
        x_values = [str(value) for value in numeric_values]
        x_positions = np.asarray(numeric_values, dtype=float)
        x_ticklabels = [str(value) for value in numeric_values]
    model_order = {
        "full_rbf": 0,
        "proposed": 0,
        "rbf": 0,
        "xgboost": 1,
        "lightgbm": 2,
        "logreg": 3,
        "svm": 4,
        "random_forest": 5,
        "mlp": 6,
        "cnn1d": 7,
        "tcn": 8,
        "autoformer": 9,
        "transformer": 10,
        "itransformer": 11,
        "no_rbf": 12,
        "no_kan_fusion": 13,
        "static_prototype": 14,
        "no_condition_input": 15,
    }
    display_names = {
        "full_rbf": "CAPT-UniShape",
        "proposed": "CAPT-UniShape",
        "rbf": "RBF",
        "xgboost": "XGBoost",
        "lightgbm": "LightGBM",
        "logreg": "LR",
        "svm": "SVM",
        "random_forest": "RF",
        "mlp": "MLP",
        "cnn1d": "CNN",
        "tcn": "TCN",
        "autoformer": "Autoformer",
        "transformer": "Transformer",
        "itransformer": "iTransformer",
        "no_rbf": "No-RBF",
        "no_kan_fusion": "No-KAN-Fusion",
        "static_prototype": "Static-Prototype",
        "no_condition_input": "No-Condition",
    }
    models = sorted({row[model_key] for row in rows}, key=lambda item: model_order.get(item, 99))
    colors = ["#c8d400", "#ef3b4a", "#31c44f", "#c24db6", "#5ab5e8", "#f47b36", "#9bdbe0", "#2b8c8c", "#c66a2e", "#b78a8e", "#6b5fd3", "#e2d61a"]

    fig, axes = plt.subplots(
        3,
        1,
        figsize=figsize_full(height_cm=17.0),
        sharex=True,
        gridspec_kw={"hspace": STACKED_SUBPLOT_ONE_LINE_HSPACE},
    )
    for ax, (panel_tag, metric_col, ylabel, y_floor, nominal_y_max) in zip(axes, METRIC_SPECS):
        metric_values = [_safe_metric_float(row, metric_col) * 100.0 for row in rows]
        proposed_rows = [row for row in rows if row.get(model_key) == "proposed"]
        proposed_values = [_safe_metric_float(row, metric_col) * 100.0 for row in proposed_rows]
        y_min, y_max, y_ticks = _noise_axis_limits_and_ticks(metric_values, proposed_values, nominal_y_max)
        for model in models:
            values: list[float] = []
            for x_value in x_values:
                if x_key == "noise_std":
                    matched = [
                        row
                        for row in rows
                        if row[model_key] == model and abs(_safe_float(row, x_key) - float(x_value)) < 1e-12
                    ]
                else:
                    matched = [row for row in rows if row[model_key] == model and str(row[x_key]) == str(x_value)]
                values.append(_safe_metric_float(matched[-1], metric_col) * 100.0 if matched else np.nan)
            is_proposed = model == "proposed"
            ax.plot(
                x_positions,
                values,
                marker=MODEL_MARKERS.get(model, "o"),
                markersize=5.1 if is_proposed else 4.0,
                linewidth=2.1 if is_proposed else 1.15,
                markerfacecolor=MODEL_COLOR_MAP.get(model, "black") if is_proposed else "white",
                markeredgecolor=MODEL_COLOR_MAP.get(model, "black"),
                markeredgewidth=0.9,
                label=display_names.get(model, model),
                color=MODEL_COLOR_MAP.get(model, colors[models.index(model) % len(colors)]),
                zorder=4 if is_proposed else 3,
            )
            if is_proposed:
                for xi, yi in zip(x_positions, values):
                    if np.isnan(yi):
                        continue
                    text_ha = "left" if xi == x_positions.min() else "right" if xi == x_positions.max() else "center"
                    ax.text(
                        xi,
                        inside_text_y(float(yi), y_min, y_max),
                        f"{yi:.2f}",
                        ha=text_ha,
                        va="bottom",
                        fontsize=ANNOTATION_SIZE,
                        color="black",
                        fontweight="bold",
                        clip_on=True,
                    )
        ax.set_ylabel(ylabel)
        ax.set_ylim(y_min, y_max)
        ax.set_yticks(y_ticks)
        style_axes(ax, grid_axis="y")
        add_panel_tag(ax, panel_tag)
        show_shared_x_axis(ax, x_positions, x_ticklabels, x_label)

    handles, labels = axes[0].get_legend_handles_labels()
    legend = axes[0].legend(
        handles,
        labels,
        ncol=4,
        loc="lower left",
        bbox_to_anchor=(0.0, 1.18, 1.0, 0.06),
        bbox_transform=axes[0].transAxes,
        mode="expand",
        handlelength=1.6,
        columnspacing=1.0,
        borderaxespad=0,
        **LEGEND_KWARGS,
    )
    style_legend_frame(legend)
    add_bottom_caption(fig, "(a) Accuracy; (b) Macro-F1; (c) Weighted-F1.")
    fig.subplots_adjust(left=0.10, right=0.98, top=0.84, bottom=0.16, hspace=STACKED_SUBPLOT_ONE_LINE_HSPACE)
    output_dir.mkdir(parents=True, exist_ok=True)
    _write_source_csv(output_dir / "official_noise_summary_source.csv", rows)
    path = output_dir / "official_noise_summary.png"
    save_paper_figure(fig, str(path.with_suffix("")), formats=PAPER_OUTPUT_FORMATS)
    plt.close(fig)
    return path


def plot_ablation_summary(summary_path: Path, output_dir: Path) -> Path:
    rows = _read_summary_csv(summary_path)
    rows = sorted(rows, key=lambda row: _safe_float(row, "test_macro_f1"), reverse=True)
    names = [row["variant"] for row in rows]
    macro_f1 = [_safe_float(row, "test_macro_f1") for row in rows]
    accuracy = [_safe_float(row, "test_accuracy") for row in rows]
    y = np.arange(len(names))

    fig, ax = plt.subplots(figsize=(8.8, max(5.2, len(names) * 0.48)))
    ax.barh(y - 0.18, macro_f1, 0.36, label="宏平均F1", color="#e15759")
    ax.barh(y + 0.18, accuracy, 0.36, label="准确率", color="#4e79a7")
    for index, value in enumerate(macro_f1):
        ax.text(value + 0.012, index - 0.18, f"{value:.3f}", va="center", fontsize=9)
    for index, value in enumerate(accuracy):
        ax.text(value + 0.012, index + 0.18, f"{value:.3f}", va="center", fontsize=9)
    ax.set_yticks(y)
    ax.set_yticklabels(names)
    ax.invert_yaxis()
    ax.set_xlim(0, 1.08)
    ax.set_xlabel("测试集得分")
    ax.set_title("官方 CAPT-UniShape 消融实验汇总")
    ax.legend()
    fig.tight_layout()
    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / "official_ablation_summary.png"
    fig.savefig(path, dpi=300, bbox_inches="tight")
    plt.close(fig)
    return path


def main() -> None:
    parser = argparse.ArgumentParser(description="绘制官方 UniShape 版 CAPT-UniShape 结果图")
    parser.add_argument("--results", nargs="+", default=[], help="一个或多个包含 metrics.json 的结果目录")
    parser.add_argument("--baseline-summary", help="基准模型对比实验输出的 summary.csv")
    parser.add_argument("--ablation-summary", help="消融实验输出的 summary.csv")
    parser.add_argument("--noise-summary", help="噪声鲁棒性实验输出的 summary.csv")
    parser.add_argument("--results-workbook", default=str(DEFAULT_RESULTS_WORKBOOK), help="用于补充对比模型指标的实验结果总表 xlsx；传空字符串则不合并")
    parser.add_argument("--output-dir", default="figures/official_unishape")
    args = parser.parse_args()

    result_dirs = [Path(item) for item in args.results]
    output_dir = Path(args.output_dir)
    results_workbook = Path(args.results_workbook) if args.results_workbook else None
    written: list[Path] = [plot_confusion_matrix(path, output_dir) for path in result_dirs]
    if len(result_dirs) >= 2:
        written.append(plot_model_comparison(result_dirs, output_dir))
    if args.baseline_summary:
        written.append(plot_baseline_summary(Path(args.baseline_summary), output_dir, results_workbook=results_workbook))
    if args.ablation_summary:
        written.append(plot_ablation_summary(Path(args.ablation_summary), output_dir))
    if args.noise_summary:
        written.append(plot_noise_summary(Path(args.noise_summary), output_dir, results_workbook=results_workbook))
    if not written:
        parser.error("请至少提供 --results、--baseline-summary、--ablation-summary 或 --noise-summary 其中之一")
    for path in written:
        print(path)


if __name__ == "__main__":
    main()
