from __future__ import annotations

import csv
import math
import unittest
from pathlib import Path

import joblib
import numpy as np
from openpyxl import Workbook, load_workbook
from sklearn.feature_selection import VarianceThreshold
from sklearn.ensemble import RandomForestClassifier
from sklearn.pipeline import Pipeline
from sklearn.svm import SVC

from experiments import refresh_snr_noise_results as snr


class SnrNoiseRefreshTests(unittest.TestCase):
    def test_add_snr_noise_hits_requested_level_for_targets_only(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            clean_path = tmp_path / "clean.npz"
            noisy_path = tmp_path / "snr_20dB.npz"
            x_op = np.ones((4, 2, 8), dtype=np.float32) * 2.0
            x_eis = np.ones((4, 1, 6), dtype=np.float32) * 3.0
            x_cond = np.ones((4, 3), dtype=np.float32) * 4.0
            split = np.asarray([0, 1, 2, 2], dtype=np.int64)
            labels = np.asarray([0, 1, 1, 2], dtype=np.int64)
            np.savez_compressed(clean_path, x_op=x_op, x_eis=x_eis, x_cond=x_cond, split=split, labels=labels)

            summary = snr.write_test_subset_with_snr_noise(
                clean_path,
                noisy_path,
                snr_db=20.0,
                noise_targets=["x_op", "x_eis"],
                seed=123,
            )

            with np.load(noisy_path) as noisy:
                self.assertEqual(noisy["x_op"].shape[0], 2)
                self.assertEqual(noisy["split"].tolist(), [2, 2])
                self.assertTrue(np.allclose(noisy["x_cond"], x_cond[split == 2]))
            self.assertEqual(set(summary["target_actual_snr_db"]), {"x_op", "x_eis"})
            for actual_snr in summary["target_actual_snr_db"].values():
                self.assertTrue(math.isclose(actual_snr, 20.0, abs_tol=1e-4))

    def test_per_sample_modality_snr_hits_requested_level_for_each_target_sample(self) -> None:
        import tempfile

        def actual_snr(clean: np.ndarray, noisy: np.ndarray) -> float:
            signal_power = float(np.mean(np.square(clean.astype(np.float64))))
            noise_power = float(np.mean(np.square((noisy - clean).astype(np.float64))))
            return float(10.0 * math.log10(signal_power / noise_power))

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = Path(tmpdir)
            clean_path = tmp_path / "clean.npz"
            noisy_path = tmp_path / "snr_15dB.npz"
            x_op = np.stack(
                [
                    np.ones((2, 4), dtype=np.float32),
                    np.ones((2, 4), dtype=np.float32) * 2.0,
                    np.ones((2, 4), dtype=np.float32) * 1.5,
                    np.ones((2, 4), dtype=np.float32) * 9.0,
                ]
            )
            x_eis = np.stack(
                [
                    np.ones((1, 5), dtype=np.float32),
                    np.ones((1, 5), dtype=np.float32) * 3.0,
                    np.ones((1, 5), dtype=np.float32) * 2.5,
                    np.ones((1, 5), dtype=np.float32) * 12.0,
                ]
            )
            x_cond = np.asarray(
                [
                    [1.0, 2.0, 3.0],
                    [2.0, 3.0, 4.0],
                    [1.0, 2.0, 4.0],
                    [10.0, 20.0, 40.0],
                ],
                dtype=np.float32,
            )
            split = np.asarray([0, 1, 2, 2], dtype=np.int64)
            labels = np.asarray([0, 1, 1, 2], dtype=np.int64)
            np.savez_compressed(clean_path, x_op=x_op, x_eis=x_eis, x_cond=x_cond, split=split, labels=labels)

            summary = snr.write_test_subset_with_snr_noise(
                clean_path,
                noisy_path,
                snr_db=15.0,
                noise_targets=["x_op", "x_eis", "x_cond"],
                seed=123,
                snr_scope="per_sample_modality",
            )

            test_mask = split == 2
            with np.load(noisy_path) as noisy:
                for key in ["x_op", "x_eis", "x_cond"]:
                    clean_subset = locals()[key][test_mask]
                    for sample_index in range(clean_subset.shape[0]):
                        self.assertTrue(
                            math.isclose(actual_snr(clean_subset[sample_index], noisy[key][sample_index]), 15.0, abs_tol=1e-3),
                            msg=f"{key} sample {sample_index} SNR mismatch",
                        )

            self.assertEqual(summary["snr_scope"], "per_sample_modality")
            self.assertEqual(len(summary["target_sample_actual_snr_db"]["x_cond"]), 2)

    def test_merge_summary_preserves_clean_rows_and_replaces_noise(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            old_path = Path(tmpdir) / "old.csv"
            fieldnames = [
                "ratio",
                "model",
                "category",
                "snr_db",
                "test_accuracy",
                "metrics_path",
            ]
            with old_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(handle, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerow({"ratio": "8:2", "model": "mlp", "category": "deep_learning", "snr_db": "clean", "test_accuracy": "0.9", "metrics_path": "clean"})
                writer.writerow({"ratio": "8:2", "model": "mlp", "category": "deep_learning", "snr_db": "30", "test_accuracy": "0.1", "metrics_path": "old_noise"})

            merged = snr.merge_clean_rows_with_noise_rows(
                old_summary_path=old_path,
                noise_rows=[
                    {
                        "ratio": "8:2",
                        "model": "mlp",
                        "category": "deep_learning",
                        "snr_db": "40",
                        "test_accuracy": "0.8",
                        "metrics_path": "new_noise",
                    }
                ],
                model_order=["mlp"],
                snr_order=["clean", "40"],
            )

            self.assertEqual([row["snr_db"] for row in merged], ["clean", "40"])
            self.assertEqual(merged[0]["metrics_path"], "clean")
            self.assertEqual(merged[1]["metrics_path"], "new_noise")

    def test_merge_can_use_same_run_clean_rows_instead_of_old_summary(self) -> None:
        merged = snr.merge_clean_rows_with_noise_rows(
            clean_rows_by_model={
                "mlp": {
                    "ratio": "8:2",
                    "model": "mlp",
                    "category": "deep_learning",
                    "snr_db": "clean",
                    "test_accuracy": "1.0",
                    "metrics_path": "same_run_clean",
                }
            },
            noise_rows=[
                {
                    "ratio": "8:2",
                    "model": "mlp",
                    "category": "deep_learning",
                    "snr_db": "40",
                    "test_accuracy": "1.0",
                    "metrics_path": "new_noise",
                }
            ],
            model_order=["mlp"],
            snr_order=["clean", "40"],
        )

        self.assertEqual(merged[0]["metrics_path"], "same_run_clean")
        self.assertEqual(float(merged[0]["test_accuracy"]), 1.0)

    def test_cli_defaults_to_comparison_results_clean_source(self) -> None:
        args = snr.parse_args([])
        self.assertEqual(args.clean_source, "comparison-results")
        self.assertEqual(args.summary_path, "results\\噪声对齐论文实验新表.csv")
        self.assertEqual(args.snr_scope, "per_sample_modality")
        self.assertEqual(args.noise_repeats, 1)
        self.assertEqual(args.baseline_profile, "noise_moderate")

    def test_noise_moderate_profile_uses_window_level_reduced_baselines(self) -> None:
        xgboost = snr.apply_baseline_profile(
            "xgboost",
            {
                "setting_name": "segment_xgboost_eis_cond_pca4",
                "input_protocol": "segment_level_non_window",
                "feature_scope": "x_eis+x_cond",
                "pca_components": 4,
                "n_estimators": 80,
                "max_depth": 2,
            },
            "noise_moderate",
        )
        lightgbm = snr.apply_baseline_profile(
            "lightgbm",
            {
                "setting_name": "segment_lightgbm_eis_cond_pca4_stump",
                "input_protocol": "segment_level_non_window",
                "feature_scope": "x_eis+x_cond",
                "pca_components": 4,
                "n_estimators": 40,
                "max_depth": 1,
                "num_leaves": 3,
            },
            "noise_moderate",
        )
        mlp = snr.apply_baseline_profile(
            "mlp",
            {
                "feature_scope": "x_cond_only",
                "hidden_dim": 12,
                "dropout": 0.5,
                "epochs": 18,
            },
            "noise_moderate",
        )

        self.assertNotIn("input_protocol", xgboost)
        self.assertEqual(xgboost["feature_scope"], "x_eis+x_cond")
        self.assertIsNone(xgboost["pca_components"])
        self.assertEqual(int(xgboost["n_estimators"]), 40)
        self.assertEqual(int(xgboost["max_depth"]), 2)

        self.assertNotIn("input_protocol", lightgbm)
        self.assertEqual(lightgbm["feature_scope"], "x_op+x_eis")
        self.assertIsNone(lightgbm["pca_components"])
        self.assertEqual(int(lightgbm["n_estimators"]), 40)
        self.assertEqual(int(lightgbm["num_leaves"]), 5)

        self.assertEqual(mlp["feature_scope"], "x_op+x_cond")
        self.assertEqual(int(mlp["hidden_dim"]), 20)
        self.assertAlmostEqual(float(mlp["dropout"]), 0.5)
        self.assertEqual(int(mlp["epochs"]), 20)
        self.assertEqual(mlp["class_weighting"], "none")

        tcn = snr.apply_baseline_profile(
            "tcn",
            {
                "feature_scope": "x_op+x_cond",
                "hidden_dim": 6,
                "dropout": 0.6,
                "epochs": 7,
            },
            "noise_moderate",
        )
        transformer = snr.apply_baseline_profile(
            "transformer",
            {
                "d_model": 24,
                "num_layers": 1,
                "dropout": 0.35,
                "epochs": 20,
            },
            "noise_moderate",
        )

        self.assertEqual(tcn["feature_scope"], "all_modalities")
        self.assertEqual(int(tcn["hidden_dim"]), 16)
        self.assertAlmostEqual(float(tcn["dropout"]), 0.4)
        self.assertEqual(int(tcn["epochs"]), 20)

        self.assertEqual(transformer["feature_scope"], "x_op+x_cond")
        self.assertEqual(int(transformer["d_model"]), 16)
        self.assertAlmostEqual(float(transformer["dropout"]), 0.45)
        self.assertEqual(int(transformer["epochs"]), 18)

    def test_compact_summary_uses_required_headers_and_four_decimals(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            out_path = Path(tmpdir) / "summary.csv"
            snr.write_compact_summary(
                out_path,
                [
                    {
                        "model": "mlp",
                        "snr_db": "40",
                        "test_accuracy": 0.987654,
                        "test_accuracy_std": 0.012345,
                        "accuracy_drop": -0.012345,
                        "test_macro_f1": 0.876543,
                        "test_macro_f1_std": 0.023456,
                        "macro_f1_drop": 0.0,
                        "test_weighted_f1": 0.765432,
                        "test_weighted_f1_std": 0.034567,
                        "weighted_f1_drop": 0.111111,
                        "data_path": "data.npz",
                        "metrics_path": "metrics.json",
                        "clean_alignment_source": "best.ckpt",
                    }
                ],
            )

            with out_path.open("r", newline="", encoding="utf-8") as handle:
                reader = csv.DictReader(handle)
                self.assertEqual(reader.fieldnames, snr.COMPACT_SUMMARY_FIELDNAMES)
                self.assertFalse(any("std" in field for field in reader.fieldnames))
                rows = list(reader)

            self.assertEqual(rows[0]["accuracy"], "0.9877")
            self.assertEqual(rows[0]["accuracy_drop"], "-0.0123")
            self.assertEqual(rows[0]["macro_f1"], "0.8765")
            self.assertEqual(rows[0]["weighted_f1_drop"], "0.1111")

    def test_row_from_metrics_does_not_clip_noisy_better_than_clean(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            metrics_path = Path(tmpdir) / "metrics.json"
            metrics_path.write_text(
                """{
  "test": {
    "accuracy": 0.95,
    "macro_f1": 0.94,
    "weighted_f1": 0.93,
    "inference_time_per_sample_ms": 1.0,
    "per_class_f1": []
  },
  "parameter_count": 1
}""",
                encoding="utf-8",
            )
            row = snr._row_from_metrics(
                model_key="mlp",
                snr_db=40.0,
                actual_snr_db_mean=40.0,
                noise_targets=["x_op"],
                data_path=Path("snr_40dB.npz"),
                metrics_path=metrics_path,
                clean_row={"test_accuracy": 0.90, "test_macro_f1": 0.89, "test_weighted_f1": 0.88},
                clean_alignment_source="same_checkpoint",
            )

            self.assertAlmostEqual(row["accuracy_drop"], -0.05)
            self.assertAlmostEqual(row["macro_f1_drop"], -0.05)
            self.assertAlmostEqual(row["weighted_f1_drop"], -0.05)

    def test_merge_can_update_selected_models_while_preserving_others(self) -> None:
        existing_rows = [
            {
                "ratio": "8:2",
                "model": "mlp",
                "category": "deep_learning",
                "snr_db": "clean",
                "test_accuracy": "0.90",
                "metrics_path": "mlp_clean",
            },
            {
                "ratio": "8:2",
                "model": "proposed",
                "category": "proposed",
                "snr_db": "clean",
                "test_accuracy": "1.00",
                "metrics_path": "proposed_clean_old",
            },
            {
                "ratio": "8:2",
                "model": "mlp",
                "category": "deep_learning",
                "snr_db": "40",
                "test_accuracy": "0.80",
                "metrics_path": "mlp_40_old",
            },
            {
                "ratio": "8:2",
                "model": "proposed",
                "category": "proposed",
                "snr_db": "40",
                "test_accuracy": "0.70",
                "metrics_path": "proposed_40_old",
            },
        ]

        merged = snr.merge_updated_rows_with_existing_rows(
            existing_rows=existing_rows,
            updated_rows=[
                {
                    "ratio": "8:2",
                    "model": "proposed",
                    "category": "proposed",
                    "snr_db": "clean",
                    "test_accuracy": "1.00",
                    "metrics_path": "proposed_clean_old",
                },
                {
                    "ratio": "8:2",
                    "model": "proposed",
                    "category": "proposed",
                    "snr_db": "40",
                    "test_accuracy": "0.88",
                    "metrics_path": "proposed_40_new",
                },
            ],
            model_order=["mlp", "proposed"],
            snr_order=["clean", "40"],
        )

        self.assertEqual(
            [(row["model"], row["snr_db"], row["metrics_path"]) for row in merged],
            [
                ("mlp", "clean", "mlp_clean"),
                ("proposed", "clean", "proposed_clean_old"),
                ("mlp", "40", "mlp_40_old"),
                ("proposed", "40", "proposed_40_new"),
            ],
        )

    def test_parse_model_path_overrides_and_resolve_reuse_artifacts(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            model_dir = Path(tmpdir) / "mlp_legacy"
            model_dir.mkdir(parents=True, exist_ok=True)
            metrics_path = model_dir / "metrics.json"
            checkpoint_path = model_dir / "best.ckpt"
            metrics_path.write_text("{}", encoding="utf-8")
            checkpoint_path.write_bytes(b"checkpoint")

            overrides = snr.parse_model_path_overrides([f"mlp={model_dir}"])
            self.assertEqual(overrides["mlp"], model_dir)

            resolved_metrics, resolved_checkpoint = snr.resolve_reuse_clean_artifacts("mlp", model_dir)
            self.assertEqual(resolved_metrics, metrics_path)
            self.assertEqual(resolved_checkpoint, checkpoint_path)

    def test_load_comparison_ml_model_artifact_requires_persisted_joblib(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            with self.assertRaises(FileNotFoundError):
                snr.load_comparison_ml_model_artifact(root, "svm")

            model_dir = root / "svm"
            model_dir.mkdir(parents=True)
            model_path = model_dir / "model.joblib"
            joblib.dump({"model": "sentinel"}, model_path)

            loaded = snr.load_comparison_ml_model_artifact(root, "svm")
            self.assertEqual(loaded["model"], "sentinel")

    def test_resolve_comparison_torch_checkpoint_requires_best_ckpt(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            with self.assertRaises(FileNotFoundError):
                snr.resolve_comparison_torch_checkpoint(root, "transformer")

            model_dir = root / "transformer"
            model_dir.mkdir(parents=True)
            checkpoint = model_dir / "best.ckpt"
            checkpoint.write_bytes(b"ckpt")

            self.assertEqual(snr.resolve_comparison_torch_checkpoint(root, "transformer"), checkpoint)

    def test_comparison_torch_settings_preserves_comparison_yaml_values(self) -> None:
        settings = snr.comparison_torch_settings({"hidden_dim": 8, "d_model": 8, "num_layers": 1, "dropout": 0.6})

        self.assertEqual(settings["hidden_dim"], 8)
        self.assertEqual(settings["d_model"], 8)
        self.assertEqual(settings["num_layers"], 1)
        self.assertEqual(settings["dropout"], 0.6)

    def test_noise_seed_matches_cli_example(self) -> None:
        self.assertEqual(snr.noise_seed_for_snr(44, 20.0), 44020000)
        self.assertEqual(snr.noise_seed_for_snr(44, 35.0), 44035000)

    def test_load_reference_clean_rows_from_test_summary(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            metrics_path = tmp / "transformer_metrics.json"
            metrics_path.write_text(
                """{
  "test": {
    "accuracy": 0.9454545454545454,
    "macro_f1": 0.9328756674294431,
    "weighted_f1": 0.9485749947992511,
    "inference_time_per_sample_ms": 0.5315909090910509,
    "per_class_f1": [{"class_id": 0, "precision": 1.0, "recall": 0.875, "f1": 0.9333333333333333}]
  },
  "parameter_count": 109763
}""",
                encoding="utf-8",
            )
            summary_path = tmp / "test_summary.csv"
            with summary_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(
                    handle,
                    fieldnames=[
                        "ratio",
                        "model",
                        "category",
                        "test_accuracy",
                        "test_macro_f1",
                        "test_weighted_f1",
                        "test_inference_ms",
                        "parameter_count",
                        "metrics_path",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "ratio": "8:2",
                        "model": "transformer",
                        "category": "transformer",
                        "test_accuracy": "0.9454545454545454",
                        "test_macro_f1": "0.9328756674294431",
                        "test_weighted_f1": "0.9485749947992511",
                        "test_inference_ms": "0.5315909090910509",
                        "parameter_count": "109763",
                        "metrics_path": str(metrics_path),
                    }
                )

            rows = snr.load_reference_clean_rows(summary_path, ["transformer"], tmp / "clean.npz")
            self.assertEqual(rows["transformer"]["snr_db"], "clean")
            self.assertEqual(rows["transformer"]["metrics_path"], str(metrics_path))
            self.assertAlmostEqual(float(rows["transformer"]["test_accuracy"]), 0.9454545454545454)

    def test_workbook_rows_format_display_names_and_percentages(self) -> None:
        row = {
            "model": "proposed",
            "snr_db": "40",
            "test_accuracy": 0.875,
            "accuracy_drop": 0.125,
            "test_macro_f1": 0.75,
            "macro_f1_drop": 0.25,
            "test_weighted_f1": 0.8,
            "test_inference_ms": 12.3456,
        }

        self.assertEqual(
            snr.summary_row_to_workbook_values(row),
            [
                "所提模型",
                "40",
                "87.50%",
                "12.50%",
                "75.00%",
                "25.00%",
                "80.00%",
                "12.346",
            ],
        )

    def test_summary_fieldnames_do_not_include_class0_columns(self) -> None:
        self.assertNotIn("class0_precision", snr.SUMMARY_FIELDNAMES)
        self.assertNotIn("class0_recall", snr.SUMMARY_FIELDNAMES)
        self.assertNotIn("class0_f1", snr.SUMMARY_FIELDNAMES)
        self.assertEqual(
            snr.EXPORT_FIELDNAMES,
            [
                "model",
                "snr_db",
                "test_accuracy",
                "accuracy_drop",
                "test_macro_f1",
                "macro_f1_drop",
                "test_weighted_f1",
                "weighted_f1_drop",
                "test_inference_ms",
                "parameter_count",
                "data_path",
                "metrics_path",
                "alignment_source",
            ],
        )

    def test_noise_repeat_plan_and_aggregation_keep_clean_single_and_summarize_noisy_rows(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            jobs = snr.build_snr_noise_jobs(
                data_root=tmp / "data",
                snr_dbs=[20.0],
                base_seed=44,
                noise_repeats=3,
            )

            self.assertEqual([job["snr_label"] for job in jobs], ["20", "20", "20"])
            self.assertEqual(len({job["npz_path"] for job in jobs}), 3)
            self.assertNotIn("clean", {Path(job["npz_path"]).stem for job in jobs})

            aggregate_dir = tmp / "aggregate"
            rows = snr.aggregate_noise_repeat_rows(
                [
                    {
                        "ratio": "8:2",
                        "model": "mlp",
                        "category": "deep_learning",
                        "snr_db": "20",
                        "snr_db_numeric": 20.0,
                        "actual_snr_db_mean": 20.0,
                        "noise_targets": "x_op+x_eis+x_cond",
                        "test_accuracy": 0.80,
                        "accuracy_drop": 0.10,
                        "test_macro_f1": 0.70,
                        "macro_f1_drop": 0.10,
                        "test_weighted_f1": 0.75,
                        "weighted_f1_drop": 0.10,
                        "test_inference_ms": 1.0,
                        "parameter_count": 10,
                        "data_path": "repeat_1.npz",
                        "metrics_path": "repeat_1.json",
                        "clean_alignment_source": "clean.json",
                    },
                    {
                        "ratio": "8:2",
                        "model": "mlp",
                        "category": "deep_learning",
                        "snr_db": "20",
                        "snr_db_numeric": 20.0,
                        "actual_snr_db_mean": 20.2,
                        "noise_targets": "x_op+x_eis+x_cond",
                        "test_accuracy": 0.90,
                        "accuracy_drop": 0.00,
                        "test_macro_f1": 0.80,
                        "macro_f1_drop": 0.00,
                        "test_weighted_f1": 0.85,
                        "weighted_f1_drop": 0.00,
                        "test_inference_ms": 2.0,
                        "parameter_count": 10,
                        "data_path": "repeat_2.npz",
                        "metrics_path": "repeat_2.json",
                        "clean_alignment_source": "clean.json",
                    },
                    {
                        "ratio": "8:2",
                        "model": "mlp",
                        "category": "deep_learning",
                        "snr_db": "20",
                        "snr_db_numeric": 20.0,
                        "actual_snr_db_mean": 19.8,
                        "noise_targets": "x_op+x_eis+x_cond",
                        "test_accuracy": 1.00,
                        "accuracy_drop": -0.10,
                        "test_macro_f1": 0.90,
                        "macro_f1_drop": -0.10,
                        "test_weighted_f1": 0.95,
                        "weighted_f1_drop": -0.10,
                        "test_inference_ms": 3.0,
                        "parameter_count": 10,
                        "data_path": "repeat_3.npz",
                        "metrics_path": "repeat_3.json",
                        "clean_alignment_source": "clean.json",
                    },
                ],
                aggregate_metrics_root=aggregate_dir,
            )

            self.assertEqual(len(rows), 1)
            row = rows[0]
            self.assertAlmostEqual(row["test_accuracy"], 0.90)
            self.assertAlmostEqual(row["test_accuracy_std"], float(np.std([0.80, 0.90, 1.00], ddof=0)))
            self.assertAlmostEqual(row["test_macro_f1"], 0.80)
            self.assertAlmostEqual(row["test_weighted_f1"], 0.85)
            self.assertTrue(Path(row["metrics_path"]).exists())

            payload = snr._metric_payload(Path(row["metrics_path"]))
            self.assertAlmostEqual(payload["test"]["accuracy"], row["test_accuracy"])
            self.assertAlmostEqual(payload["test"]["macro_f1"], row["test_macro_f1"])
            self.assertAlmostEqual(payload["test"]["weighted_f1"], row["test_weighted_f1"])
            self.assertAlmostEqual(payload["test"]["std"]["accuracy"], row["test_accuracy_std"])
            self.assertAlmostEqual(payload["test"]["std"]["macro_f1"], row["test_macro_f1_std"])
            self.assertAlmostEqual(payload["test"]["std"]["weighted_f1"], row["test_weighted_f1_std"])
            self.assertAlmostEqual(payload["test"]["std"]["inference_time_per_sample_ms"], row["test_inference_ms_std"])

            merged = snr.merge_clean_rows_with_noise_rows(
                clean_rows_by_model={
                    "mlp": {
                        "ratio": "8:2",
                        "model": "mlp",
                        "category": "deep_learning",
                        "snr_db": "clean",
                        "test_accuracy": 1.0,
                        "metrics_path": "clean.json",
                    }
                },
                noise_rows=rows,
                model_order=["mlp"],
                snr_order=["clean", "20"],
            )
            self.assertEqual([row["snr_db"] for row in merged], ["clean", "20"])

    def test_update_workbook_snr_sheet_uses_times_new_roman_font(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "snr.xlsx"
            workbook = Workbook()
            workbook.active.title = "Sheet1"
            workbook.save(workbook_path)
            workbook.close()

            snr.update_workbook_snr_sheet(
                workbook_path,
                [
                    {
                        "model": "proposed",
                        "snr_db": "40",
                        "test_accuracy": 1.0,
                        "accuracy_drop": 0.0,
                        "test_macro_f1": 1.0,
                        "macro_f1_drop": 0.0,
                        "test_weighted_f1": 1.0,
                        "test_inference_ms": 20.1142,
                    }
                ],
            )

            saved = load_workbook(workbook_path, read_only=True)
            try:
                worksheet = saved["SNR噪声对比"]
                self.assertEqual(worksheet["A1"].font.name, "Times New Roman")
                self.assertEqual(worksheet["A2"].font.name, "Times New Roman")
                self.assertEqual(worksheet["H2"].font.name, "Times New Roman")
            finally:
                saved.close()
                del saved
                import gc

                gc.collect()

    def test_update_workbook_snr_sheet_removes_legacy_class0_column(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = Path(tmpdir) / "snr.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "SNR噪声对比"
            worksheet.append(["模型", "SNR(dB)", "Accuracy", "Acc下降", "Macro-F1", "Macro下降", "Weighted-F1", "第0类 Recall", "推理时间(ms/sample)"])
            worksheet.append(["旧模型", "clean", "100.00%", "0.00%", "100.00%", "0.00%", "100.00%", "100.00%", "1.000"])
            workbook.save(workbook_path)
            workbook.close()

            snr.update_workbook_snr_sheet(
                workbook_path,
                [
                    {
                        "model": "itransformer",
                        "snr_db": "clean",
                        "test_accuracy": 0.8689,
                        "accuracy_drop": 0.0,
                        "test_macro_f1": 0.6105,
                        "macro_f1_drop": 0.0,
                        "test_weighted_f1": 0.8084,
                        "test_inference_ms": 0.1,
                    }
                ],
            )

            saved = load_workbook(workbook_path, read_only=True)
            try:
                worksheet = saved["SNR噪声对比"]
                headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
                values = [worksheet.cell(2, column).value for column in range(1, worksheet.max_column + 1)]
            finally:
                saved.close()
                del saved
                import gc

                gc.collect()

            self.assertEqual(
                headers,
                ["模型", "SNR(dB)", "Accuracy", "Acc下降", "Macro-F1", "Macro下降", "Weighted-F1", "推理时间(ms/sample)"],
            )
            self.assertEqual(len(values), 8)

    def test_export_row_uses_percent_strings_with_two_decimals(self) -> None:
        row = snr.export_summary_row(
            {
                "model": "proposed",
                "snr_db": "40",
                "test_accuracy": 1.0,
                "test_accuracy_std": 0.01,
                "accuracy_drop": 0.12345,
                "test_macro_f1": 0.815384615,
                "test_macro_f1_std": 0.02,
                "macro_f1_drop": 0.184615384,
                "test_weighted_f1": 0.931934732,
                "test_weighted_f1_std": 0.03,
                "weighted_f1_drop": 0.068065268,
                "test_inference_ms": 25.2239,
                "test_inference_ms_std": 0.04,
                "parameter_count": 6496573,
                "data_path": "clean.npz",
                "metrics_path": "metrics.json",
                "clean_alignment_source": "comparison_summary_reference",
            }
        )
        self.assertEqual(row["test_accuracy"], "100.00%")
        self.assertEqual(row["accuracy_drop"], "12.35%")
        self.assertEqual(row["test_macro_f1"], "81.54%")
        self.assertEqual(row["macro_f1_drop"], "18.46%")
        self.assertEqual(row["test_weighted_f1"], "93.19%")
        self.assertEqual(row["weighted_f1_drop"], "6.81%")
        self.assertEqual(row["test_inference_ms"], "25.2239")
        self.assertEqual(row["alignment_source"], "comparison_summary_reference")
        self.assertFalse(any("std" in field for field in row))

    def test_normalize_legacy_summary_row_schema(self) -> None:
        row = snr.normalize_summary_row_schema(
            {
                "model": "proposed",
                "snr_db": "clean",
                "accuracy": "1.0",
                "macro_f1": "1.0",
                "weighted_f1": "1.0",
                "data_path": "clean.npz",
                "metrics_path": "metrics.json",
            }
        )

        self.assertEqual(row["model"], "proposed")
        self.assertEqual(row["category"], "proposed")
        self.assertEqual(row["test_accuracy"], "1.0")
        self.assertEqual(row["test_macro_f1"], "1.0")
        self.assertEqual(row["test_weighted_f1"], "1.0")

    def test_load_model_order_from_current_comparison_config_excludes_models(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            config_path = Path(tmpdir) / "current_comparison_models.yaml"
            config_path.write_text(
                """proposed:
  model_name: proposed_model
baselines:
  xgboost: {}
  lightgbm: {}
  mlp: {}
  tcn: {}
  autoformer: {}
excluded_models:
  - lightgbm
  - lstm
""",
                encoding="utf-8",
            )

            self.assertEqual(
                snr.load_model_order_from_current_comparison_config(config_path),
                ["proposed", "xgboost", "mlp", "tcn", "autoformer"],
            )

    def test_parse_args_defaults_models_to_config_selection(self) -> None:
        args = snr.parse_args([])

        self.assertIsNone(args.models)

    def test_select_models_for_run_defaults_to_available_config_order(self) -> None:
        available = ["proposed", "xgboost", "lightgbm", "mlp", "tcn", "autoformer"]

        self.assertEqual(snr.select_models_for_run(None, available), available)
        self.assertEqual(snr.select_models_for_run(["mlp", "lstm", "lightgbm"], available), ["mlp", "lightgbm"])

    def test_load_reference_clean_rows_from_comparison_summaries(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            baseline_summary = tmp / "test_summary.csv"
            proposed_summary = tmp / "proposed_summary.csv"

            with baseline_summary.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(
                    handle,
                    fieldnames=[
                        "ratio",
                        "model",
                        "accuracy",
                        "macro_f1",
                        "weighted_f1",
                        "inference_ms",
                        "parameter_count",
                        "metrics_path",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "ratio": "8_2",
                        "model": "svm",
                        "accuracy": "89.39%",
                        "macro_f1": "58.62%",
                        "weighted_f1": "85.37%",
                        "inference_ms": "0.0182",
                        "parameter_count": "6.0000",
                        "metrics_path": "results\\baseline\\svm\\metrics.json",
                    }
                )

            with proposed_summary.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(
                    handle,
                    fieldnames=[
                        "ratio",
                        "data",
                        "output_dir",
                        "test_accuracy",
                        "test_macro_f1",
                        "test_weighted_f1",
                        "class0_recall",
                        "test_inference_ms",
                        "parameter_count",
                        "metrics_json",
                        "selected_ckpt",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "ratio": "8_2",
                        "data": "data\\processed\\self_seed44_8_2.npz",
                        "output_dir": "results\\proposed\\8_2",
                        "test_accuracy": "100.00%",
                        "test_macro_f1": "100.00%",
                        "test_weighted_f1": "100.00%",
                        "class0_recall": "100.00%",
                        "test_inference_ms": "25.2239",
                        "parameter_count": "6496573",
                        "metrics_json": "results\\proposed\\8_2\\metrics.json",
                        "selected_ckpt": "results\\proposed\\8_2\\selected.ckpt",
                    }
                )

            rows = snr.load_reference_clean_rows_from_comparison_summaries(
                baseline_summary_path=baseline_summary,
                proposed_summary_path=proposed_summary,
                model_order=["proposed", "svm"],
                data_path=tmp / "clean.npz",
            )

            self.assertAlmostEqual(float(rows["svm"]["test_accuracy"]), 0.8939)
            self.assertEqual(rows["svm"]["metrics_path"], "results\\baseline\\svm\\metrics.json")
            self.assertAlmostEqual(float(rows["proposed"]["test_macro_f1"]), 1.0)
            self.assertEqual(rows["proposed"]["metrics_path"], "results\\proposed\\8_2\\metrics.json")

    def test_load_reference_clean_rows_from_comparison_summaries_uses_requested_ratio(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            baseline_summary = tmp / "test_summary.csv"
            proposed_summary = tmp / "proposed_summary.csv"

            with baseline_summary.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(
                    handle,
                    fieldnames=[
                        "ratio",
                        "model",
                        "accuracy",
                        "macro_f1",
                        "weighted_f1",
                        "inference_ms",
                        "parameter_count",
                        "metrics_path",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "ratio": "8_2",
                        "model": "cnn1d",
                        "accuracy": "99.18%",
                        "macro_f1": "98.79%",
                        "weighted_f1": "99.19%",
                        "inference_ms": "0.1106",
                        "parameter_count": "1000",
                        "metrics_path": "results\\baseline\\cnn1d\\8_2\\metrics.json",
                    }
                )
                writer.writerow(
                    {
                        "ratio": "6_4",
                        "model": "cnn1d",
                        "accuracy": "80.33%",
                        "macro_f1": "55.00%",
                        "weighted_f1": "74.00%",
                        "inference_ms": "0.1200",
                        "parameter_count": "900",
                        "metrics_path": "results\\baseline\\cnn1d\\6_4\\metrics.json",
                    }
                )
            proposed_summary.write_text(
                "ratio,output_dir,test_accuracy,test_macro_f1,test_weighted_f1,test_inference_ms,parameter_count,metrics_json,selected_ckpt\n",
                encoding="utf-8",
            )

            rows = snr.load_reference_clean_rows_from_comparison_summaries(
                baseline_summary_path=baseline_summary,
                proposed_summary_path=proposed_summary,
                model_order=["cnn1d"],
                data_path=tmp / "clean.npz",
                ratio_key="6_4",
                ratio_label="6:4",
            )

            self.assertEqual(rows["cnn1d"]["ratio"], "6:4")
            self.assertAlmostEqual(float(rows["cnn1d"]["test_accuracy"]), 0.8033)
            self.assertEqual(rows["cnn1d"]["metrics_path"], "results\\baseline\\cnn1d\\6_4\\metrics.json")

    def test_load_clean_rows_from_comparison_result_csv_uses_requested_ratio(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            comparison_results = tmp / "对比实验结果.csv"
            with comparison_results.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(
                    handle,
                    fieldnames=[
                        "ratio",
                        "model",
                        "accuracy",
                        "macro_f1",
                        "weighted_f1",
                        "inference_ms",
                        "parameter_count",
                        "result_dir",
                        "metrics_path",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "ratio": "8_2",
                        "model": "itransformer",
                        "accuracy": "100.00%",
                        "macro_f1": "100.00%",
                        "weighted_f1": "100.00%",
                        "inference_ms": "0.1",
                        "parameter_count": "214275",
                        "result_dir": "bad",
                        "metrics_path": "bad/metrics.json",
                    }
                )
                writer.writerow(
                    {
                        "ratio": "6_4",
                        "model": "itransformer",
                        "accuracy": "86.89%",
                        "macro_f1": "60.69%",
                        "weighted_f1": "80.88%",
                        "inference_ms": "0.64",
                        "parameter_count": "3539",
                        "result_dir": "results\\baseline\\6_4\\itransformer",
                        "metrics_path": "results\\baseline\\6_4\\itransformer\\metrics.json",
                    }
                )

            rows = snr.load_clean_rows_from_comparison_result_csv(
                comparison_results_path=comparison_results,
                model_order=["itransformer"],
                data_path=tmp / "clean.npz",
                ratio_key="6_4",
                ratio_label="6:4",
            )

            self.assertEqual(rows["itransformer"]["ratio"], "6:4")
            self.assertAlmostEqual(float(rows["itransformer"]["test_accuracy"]), 0.8689)
            self.assertAlmostEqual(float(rows["itransformer"]["test_macro_f1"]), 0.6069)
            self.assertEqual(int(rows["itransformer"]["parameter_count"]), 3539)

    def test_assert_clean_rows_match_reference_rejects_wrong_itransformer(self) -> None:
        reference = {
            "itransformer": {
                "test_accuracy": 0.8689,
                "test_macro_f1": 0.6069,
                "test_weighted_f1": 0.8088,
                "parameter_count": 3539,
            }
        }
        candidate = {
            "itransformer": {
                "test_accuracy": 1.0,
                "test_macro_f1": 1.0,
                "test_weighted_f1": 1.0,
                "parameter_count": 214275,
            }
        }

        with self.assertRaisesRegex(ValueError, "itransformer"):
            snr.assert_clean_rows_match_reference(
                reference_rows=reference,
                candidate_rows=candidate,
                model_order=["itransformer"],
            )

    def test_archive_existing_summary_moves_new_table_to_old_table(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            summary_path = tmp / "噪声对齐论文实验新表.csv"
            old_path = tmp / "噪声对齐论文实验旧表.csv"
            summary_path.write_text("bad itransformer table", encoding="utf-8")
            old_path.write_text("previous old table", encoding="utf-8")

            snr.archive_existing_summary(summary_path, old_path)

            self.assertFalse(summary_path.exists())
            self.assertEqual(old_path.read_text(encoding="utf-8"), "bad itransformer table")

    def test_load_reference_clean_rows_from_comparison_summaries_derives_proposed_paths_from_output_dir(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            baseline_summary = tmp / "test_summary.csv"
            proposed_summary = tmp / "proposed_summary.csv"
            baseline_summary.write_text("ratio,model,accuracy,macro_f1,weighted_f1,inference_ms,parameter_count,metrics_path\n", encoding="utf-8")

            with proposed_summary.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(
                    handle,
                    fieldnames=[
                        "ratio",
                        "output_dir",
                        "test_accuracy",
                        "test_macro_f1",
                        "test_weighted_f1",
                        "test_inference_ms",
                        "parameter_count",
                        "metrics_json",
                        "selected_ckpt",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "ratio": "8_2",
                        "output_dir": "results\\proposed\\8_2",
                        "test_accuracy": "100.00%",
                        "test_macro_f1": "100.00%",
                        "test_weighted_f1": "100.00%",
                        "test_inference_ms": "25.2239",
                        "parameter_count": "6496573",
                        "metrics_json": "OK",
                        "selected_ckpt": "OK",
                    }
                )

            rows = snr.load_reference_clean_rows_from_comparison_summaries(
                baseline_summary_path=baseline_summary,
                proposed_summary_path=proposed_summary,
                model_order=["proposed"],
                data_path=tmp / "clean.npz",
            )

            self.assertEqual(rows["proposed"]["metrics_path"], "results\\proposed\\8_2\\metrics.json")
            self.assertTrue(rows["proposed"]["checkpoint_path"].endswith("results\\proposed\\8_2\\selected.ckpt"))

    def test_choose_clean_row_falls_back_to_recomputed_on_mismatch(self) -> None:
        reference = {
            "model": "svm",
            "test_accuracy": 0.8939,
            "test_macro_f1": 0.5862,
            "test_weighted_f1": 0.8537,
            "metrics_path": "reference_metrics.json",
        }
        recomputed = {
            "model": "svm",
            "test_accuracy": 0.8000,
            "test_macro_f1": 0.5000,
            "test_weighted_f1": 0.7000,
            "metrics_path": "recomputed_metrics.json",
        }

        chosen = snr.choose_clean_row_for_noise_alignment(reference, recomputed, tolerance=1e-6)
        self.assertEqual(chosen["metrics_path"], "recomputed_metrics.json")
        self.assertEqual(chosen["clean_alignment_source"], "recomputed_due_to_mismatch")

    def test_proposed_modality_rows_exclude_clean_and_tag_targets(self) -> None:
        rows = snr.build_proposed_modality_summary_rows(
            [
                {
                    "model": "proposed",
                    "snr_db": "40",
                    "noise_targets": "x_op",
                    "test_accuracy": 0.9,
                    "accuracy_drop": 0.1,
                    "test_macro_f1": 0.8,
                    "macro_f1_drop": 0.2,
                    "test_weighted_f1": 0.85,
                    "weighted_f1_drop": 0.15,
                    "class0_recall": 0.75,
                    "test_inference_ms": 12.3,
                    "metrics_path": "op_40.json",
                },
                {
                    "model": "proposed",
                    "snr_db": "clean",
                    "noise_targets": "x_op",
                    "test_accuracy": 1.0,
                    "accuracy_drop": 0.0,
                    "test_macro_f1": 1.0,
                    "macro_f1_drop": 0.0,
                    "test_weighted_f1": 1.0,
                    "weighted_f1_drop": 0.0,
                    "class0_recall": 1.0,
                    "test_inference_ms": 12.0,
                    "metrics_path": "op_clean.json",
                },
                {
                    "model": "proposed",
                    "snr_db": "35",
                    "noise_targets": "x_cond",
                    "test_accuracy": 0.88,
                    "accuracy_drop": 0.12,
                    "test_macro_f1": 0.77,
                    "macro_f1_drop": 0.23,
                    "test_weighted_f1": 0.83,
                    "weighted_f1_drop": 0.17,
                    "class0_recall": 0.7,
                    "test_inference_ms": 12.5,
                    "metrics_path": "cond_35.json",
                },
            ]
        )

        self.assertEqual(len(rows), 2)
        self.assertEqual([row["noise_targets"] for row in rows], ["x_op", "x_cond"])
        self.assertEqual([row["snr_db"] for row in rows], ["40", "35"])

    def test_resolve_model_run_settings_reads_current_config(self) -> None:
        payload = {
            "proposed": {
                "config_file": "configs/proposed.yaml",
                "batch_size": 32,
                "epochs": 80,
                "lr": 0.0001,
            },
            "baselines": {
                "svm": {
                    "C": 0.02,
                    "max_iter": 5000,
                },
                "cnn1d": {
                    "hidden_dim": 8,
                    "epochs": 12,
                    "ratio_overrides": {
                        "8_2": {
                            "dropout": 0.55,
                        }
                    },
                },
            },
        }

        proposed = snr.resolve_model_run_settings(payload, "proposed", ratio_key="8_2")
        cnn1d = snr.resolve_model_run_settings(payload, "cnn1d", ratio_key="8_2")
        svm = snr.resolve_model_run_settings(payload, "svm", ratio_key="8_2")

        self.assertEqual(proposed["config_file"], "configs/proposed.yaml")
        self.assertEqual(cnn1d["hidden_dim"], 8)
        self.assertEqual(cnn1d["dropout"], 0.55)
        self.assertEqual(svm["C"], 0.02)

    def test_parse_args_supports_ratio_specific_noise_runs(self) -> None:
        args = snr.parse_args(
            [
                "--ratio-key",
                "6_4",
                "--clean-npz",
                "data/processed/self_seed44_6_4.npz",
                "--output-root",
                "results/current_snr_noise_6_4_seed44_artifacts",
                "--data-root",
                "data/processed/current_snr_noise_6_4_seed44_artifacts",
            ]
        )

        self.assertEqual(args.ratio_key, "6_4")
        self.assertEqual(args.ratio_label, "6:4")

    def test_row_from_metrics_uses_requested_ratio_label(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            metrics_path = Path(tmpdir) / "metrics.json"
            metrics_path.write_text(
                """{
  "test": {
    "accuracy": 0.8,
    "macro_f1": 0.7,
    "weighted_f1": 0.75,
    "inference_time_per_sample_ms": 1.0,
    "per_class_f1": []
  },
  "parameter_count": 1
}""",
                encoding="utf-8",
            )

            row = snr._row_from_metrics(
                model_key="mlp",
                snr_db=40.0,
                actual_snr_db_mean=40.0,
                noise_targets=["x_op"],
                data_path=Path("snr_40dB.npz"),
                metrics_path=metrics_path,
                clean_row={"test_accuracy": 0.9, "test_macro_f1": 0.8, "test_weighted_f1": 0.85},
                clean_alignment_source="same_checkpoint",
                ratio_label="6:4",
            )

            self.assertEqual(row["ratio"], "6:4")

    def test_flatten_split_for_model_honors_feature_scope_setting(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            npz_path = Path(tmpdir) / "toy.npz"
            np.savez_compressed(
                npz_path,
                x_op=np.arange(3 * 2 * 4, dtype=np.float32).reshape(3, 2, 4),
                x_eis=(100 + np.arange(3 * 1 * 3, dtype=np.float32)).reshape(3, 1, 3),
                x_cond=(200 + np.arange(3 * 2, dtype=np.float32)).reshape(3, 2),
                split=np.asarray([0, 2, 2], dtype=np.int64),
                labels=np.asarray([0, 1, 2], dtype=np.int64),
            )

            x_op_only, y_op_only = snr._flatten_split_for_model(
                npz_path,
                split_value=2,
                model_key="logreg",
                settings={"feature_scope": "x_op_only"},
            )
            x_all, y_all = snr._flatten_split_for_model(
                npz_path,
                split_value=2,
                model_key="svm",
                settings={"feature_scope": "all_modalities"},
            )

            self.assertEqual(x_op_only.shape, (2, 8))
            self.assertEqual(x_all.shape, (2, 13))
            self.assertTrue(np.array_equal(y_op_only, np.asarray([1, 2], dtype=np.int64)))
            self.assertTrue(np.array_equal(y_all, np.asarray([1, 2], dtype=np.int64)))

    def test_segment_level_split_for_model_aggregates_by_group_id(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            npz_path = Path(tmpdir) / "toy.npz"
            np.savez_compressed(
                npz_path,
                x_op=np.asarray([[[1.0, 3.0]], [[5.0, 7.0]], [[10.0, 14.0]]], dtype=np.float32),
                x_eis=np.ones((3, 1, 2), dtype=np.float32),
                x_cond=np.asarray([[2.0], [4.0], [8.0]], dtype=np.float32),
                split=np.asarray([2, 2, 2], dtype=np.int64),
                group_ids=np.asarray([5, 5, 6], dtype=np.int64),
                labels=np.asarray([1, 1, 2], dtype=np.int64),
            )

            features, labels = snr.feature_split_for_model(
                npz_path,
                split_value=2,
                model_key="xgboost",
                settings={
                    "input_protocol": "segment_level_non_window",
                    "feature_scope": "x_op+x_cond",
                    "segment_statistics": ["mean", "max"],
                },
            )

            self.assertEqual(features.shape, (2, 6))
            self.assertTrue(np.array_equal(labels, np.asarray([1, 2], dtype=np.int64)))
            self.assertTrue(np.allclose(features[0], np.asarray([3.0, 5.0, 5.0, 7.0, 3.0, 4.0], dtype=np.float32)))

    def test_normalize_feature_scope_supports_eis_and_cond_aliases(self) -> None:
        self.assertEqual(snr._normalize_feature_scope("mlp", {"feature_scope": "x_eis_only"}), "x_eis_only")
        self.assertEqual(snr._normalize_feature_scope("mlp", {"feature_scope": "eis_only"}), "x_eis_only")
        self.assertEqual(snr._normalize_feature_scope("mlp", {"feature_scope": "eis"}), "x_eis_only")
        self.assertEqual(snr._normalize_feature_scope("mlp", {"feature_scope": "x_cond_only"}), "x_cond_only")
        self.assertEqual(snr._normalize_feature_scope("mlp", {"feature_scope": "cond_only"}), "x_cond_only")
        self.assertEqual(snr._normalize_feature_scope("mlp", {"feature_scope": "cond"}), "x_cond_only")
        self.assertEqual(snr._normalize_feature_scope("mlp", {"feature_scope": "x_op + x_cond"}), "x_op+x_cond")
        self.assertEqual(snr._normalize_feature_scope("mlp", {"feature_scope": "cond+eis"}), "x_eis+x_cond")

    def test_scope_torch_dataset_features_zeroes_unused_modalities(self) -> None:
        class ToyTorchDataset:
            def __init__(self) -> None:
                self.x_op = np.ones((2, 2, 4), dtype=np.float32)
                self.x_eis = np.ones((2, 1, 3), dtype=np.float32) * 2.0
                self.x_cond = np.ones((2, 3), dtype=np.float32) * 3.0
                self.labels = np.asarray([0, 1], dtype=np.int64)

            def __len__(self) -> int:
                return int(len(self.labels))

            def __getitem__(self, index: int):
                return self.x_op[index], self.x_eis[index], self.x_cond[index], self.labels[index]

        original = ToyTorchDataset()

        scoped = snr._scope_torch_dataset_features(original, "mlp", {"feature_scope": "x_op_only"})
        eis_scoped = snr._scope_torch_dataset_features(original, "mlp", {"feature_scope": "x_eis_only"})
        cond_scoped = snr._scope_torch_dataset_features(original, "mlp", {"feature_scope": "x_cond_only"})
        op_cond_scoped = snr._scope_torch_dataset_features(original, "mlp", {"feature_scope": "x_op+x_cond"})
        all_modalities = snr._scope_torch_dataset_features(original, "mlp", {"feature_scope": "all_modalities"})

        self.assertTrue(np.array_equal(scoped.x_op, original.x_op))
        self.assertTrue(np.array_equal(scoped.x_eis, np.zeros_like(original.x_eis)))
        self.assertTrue(np.array_equal(scoped.x_cond, np.zeros_like(original.x_cond)))
        self.assertTrue(np.array_equal(eis_scoped.x_op, np.zeros_like(original.x_op)))
        self.assertTrue(np.array_equal(eis_scoped.x_eis, original.x_eis))
        self.assertTrue(np.array_equal(eis_scoped.x_cond, np.zeros_like(original.x_cond)))
        self.assertTrue(np.array_equal(cond_scoped.x_op, np.zeros_like(original.x_op)))
        self.assertTrue(np.array_equal(cond_scoped.x_eis, np.zeros_like(original.x_eis)))
        self.assertTrue(np.array_equal(cond_scoped.x_cond, original.x_cond))
        self.assertTrue(np.array_equal(op_cond_scoped.x_op, original.x_op))
        self.assertTrue(np.array_equal(op_cond_scoped.x_eis, np.zeros_like(original.x_eis)))
        self.assertTrue(np.array_equal(op_cond_scoped.x_cond, original.x_cond))
        self.assertTrue(np.array_equal(all_modalities.x_eis, original.x_eis))
        self.assertTrue(np.array_equal(all_modalities.x_cond, original.x_cond))
        self.assertTrue(np.array_equal(original.x_eis, np.ones_like(original.x_eis) * 2.0))

    def test_scope_torch_npz_features_zeroes_unused_modalities_for_eis_only(self) -> None:
        import tempfile

        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)
            npz_path = tmp / "toy.npz"
            x_op = np.ones((2, 2, 4), dtype=np.float32)
            x_eis = np.ones((2, 1, 3), dtype=np.float32) * 2.0
            x_cond = np.ones((2, 3), dtype=np.float32) * 3.0
            np.savez_compressed(
                npz_path,
                x_op=x_op,
                x_eis=x_eis,
                x_cond=x_cond,
                split=np.asarray([0, 2], dtype=np.int64),
                labels=np.asarray([0, 1], dtype=np.int64),
            )

            scoped_path = snr._scope_torch_npz_features(
                npz_path,
                tmp / "scoped",
                "mlp",
                {"feature_scope": "x_eis+x_cond"},
            )

            with np.load(scoped_path) as scoped:
                self.assertTrue(np.array_equal(scoped["x_op"], np.zeros_like(x_op)))
                self.assertTrue(np.array_equal(scoped["x_eis"], x_eis))
                self.assertTrue(np.array_equal(scoped["x_cond"], x_cond))

    def test_build_ml_model_from_settings_honors_pipeline_configuration(self) -> None:
        svm_model = snr._build_ml_model_from_settings(
            "svm",
            {
                "use_scaler": True,
                "variance_threshold": 0.0001,
                "pca_components": 8,
                "kernel": "rbf",
                "C": 1.0,
                "gamma": "scale",
                "class_weighting": "balanced",
                "random_state": 44,
            },
        )
        rf_model = snr._build_ml_model_from_settings(
            "random_forest",
            {
                "use_scaler": False,
                "pca_components": None,
                "n_estimators": 160,
                "max_depth": 8,
                "min_samples_leaf": 2,
                "class_weighting": "balanced_subsample",
                "random_state": 44,
            },
        )

        self.assertIsInstance(svm_model, Pipeline)
        self.assertEqual(list(svm_model.named_steps.keys()), ["variance", "scaler", "pca", "classifier"])
        self.assertIsInstance(svm_model.named_steps["variance"], VarianceThreshold)
        self.assertAlmostEqual(float(svm_model.named_steps["variance"].threshold), 0.0001)
        self.assertIsInstance(svm_model.named_steps["classifier"], SVC)
        self.assertEqual(svm_model.named_steps["classifier"].kernel, "rbf")

        self.assertIsInstance(rf_model, RandomForestClassifier)
        self.assertEqual(rf_model.class_weight, "balanced_subsample")
        self.assertEqual(rf_model.max_depth, 8)

    def test_current_comparison_config_uses_modest_non_degenerate_settings(self) -> None:
        config = snr.load_current_comparison_config(
            Path(__file__).resolve().parents[1] / "configs" / "current_comparison_models.yaml"
        )
        baselines = config["baselines"]

        self.assertEqual(
            list(baselines.keys()),
            ["xgboost", "lightgbm", "mlp", "tcn", "autoformer", "transformer", "itransformer"],
        )

        self.assertEqual(baselines["xgboost"]["input_protocol"], "segment_level_non_window")
        self.assertEqual(baselines["xgboost"]["feature_scope"], "x_eis+x_cond")
        self.assertEqual(snr._resolve_pca_components(baselines["xgboost"].get("pca_components")), 4)
        self.assertEqual(int(baselines["xgboost"]["n_estimators"]), 80)

        self.assertEqual(baselines["lightgbm"]["input_protocol"], "segment_level_non_window")
        self.assertEqual(baselines["lightgbm"]["feature_scope"], "x_eis+x_cond")
        self.assertEqual(snr._resolve_pca_components(baselines["lightgbm"].get("pca_components")), 4)
        self.assertEqual(int(baselines["lightgbm"]["n_estimators"]), 40)

        self.assertNotIn("run_policy", baselines["mlp"])
        self.assertEqual(baselines["mlp"]["feature_scope"], "x_cond_only")
        self.assertEqual(int(baselines["mlp"]["hidden_dim"]), 12)
        self.assertAlmostEqual(float(baselines["mlp"]["dropout"]), 0.5)
        self.assertEqual(int(baselines["mlp"]["epochs"]), 18)

        self.assertEqual(baselines["tcn"]["feature_scope"], "x_op+x_cond")
        self.assertEqual(int(baselines["tcn"]["hidden_dim"]), 6)
        self.assertAlmostEqual(float(baselines["tcn"]["dropout"]), 0.6)
        self.assertEqual(int(baselines["tcn"]["epochs"]), 7)

        self.assertEqual(baselines["autoformer"]["feature_scope"], "x_eis_only")
        self.assertEqual(int(baselines["autoformer"]["d_model"]), 8)
        self.assertEqual(int(baselines["autoformer"]["num_layers"]), 1)
        self.assertEqual(int(baselines["autoformer"]["moving_avg_kernel"]), 5)

    def test_current_config_preserves_comparison_strength_for_6_4_torch_overrides(self) -> None:
        config = snr.load_current_comparison_config(
            Path(__file__).resolve().parents[1] / "configs" / "current_comparison_models.yaml"
        )

        proposed = snr.resolve_model_run_settings(config, "proposed", ratio_key="6_4")
        tcn = snr.resolve_model_run_settings(config, "tcn", ratio_key="6_4")
        autoformer = snr.resolve_model_run_settings(config, "autoformer", ratio_key="6_4")

        self.assertEqual(proposed["setting_name"], "proposed_clean_checkpoint_finetune_6_4")
        self.assertIsNone(proposed["init_checkpoint"])
        self.assertEqual(int(proposed["epochs"]), 12)
        self.assertEqual(int(proposed["patience"]), 4)
        self.assertAlmostEqual(float(proposed["lr"]), 0.00002)

        self.assertEqual(tcn["setting_name"], "compact_tcn_op_cond_small")
        self.assertEqual(tcn["feature_scope"], "x_op+x_cond")
        self.assertEqual(int(tcn["hidden_dim"]), 6)
        self.assertEqual(int(tcn["epochs"]), 7)
        self.assertEqual(str(tcn["class_weighting"]).lower(), "sqrt_balanced")

        self.assertEqual(autoformer["setting_name"], "compact_autoformer_op_cond_small")
        self.assertEqual(autoformer["feature_scope"], "x_eis_only")
        self.assertEqual(int(autoformer["d_model"]), 8)
        self.assertEqual(int(autoformer["num_layers"]), 1)
        self.assertEqual(int(autoformer["moving_avg_kernel"]), 5)

    def test_current_config_preserves_fair_tree_boosting_ml_settings(self) -> None:
        config = snr.load_current_comparison_config(
            Path(__file__).resolve().parents[1] / "configs" / "current_comparison_models.yaml"
        )

        xgboost = snr.resolve_model_run_settings(config, "xgboost", ratio_key="6_4")
        lightgbm = snr.resolve_model_run_settings(config, "lightgbm", ratio_key="6_4")

        self.assertEqual(xgboost["setting_name"], "segment_xgboost_eis_cond_pca4")
        self.assertEqual(xgboost["input_protocol"], "segment_level_non_window")
        self.assertEqual(xgboost["feature_scope"], "x_eis+x_cond")
        self.assertEqual(snr._resolve_pca_components(xgboost["pca_components"]), 4)
        self.assertEqual(int(xgboost["n_estimators"]), 80)
        self.assertEqual(int(xgboost["random_state"]), 44)

        self.assertEqual(lightgbm["setting_name"], "segment_lightgbm_eis_cond_pca4_stump")
        self.assertEqual(lightgbm["input_protocol"], "segment_level_non_window")
        self.assertEqual(lightgbm["feature_scope"], "x_eis+x_cond")
        self.assertEqual(snr._resolve_pca_components(lightgbm["pca_components"]), 4)
        self.assertEqual(int(lightgbm["n_estimators"]), 40)
        self.assertEqual(int(lightgbm["random_state"]), 44)


if __name__ == "__main__":
    unittest.main()

